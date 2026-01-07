# =============================================================================
# Sidikjari - Python-Based Metadata Extraction Tool
# =============================================================================
#
# Author: Keith Pachulski
# Company: Red Cell Security, LLC
# Email: keith@redcellsecurity.org
# Website: www.redcellsecurity.org
#
# Copyright (c) 2025 Keith Pachulski. All rights reserved.
#
# License: This software is licensed under the MIT License.
#          You are free to use, modify, and distribute this software
#          in accordance with the terms of the license.
#
# Purpose: This module is part of the Sidikjari metadata extraction system, designed to
#          analyze documents and extract valuable metadata information.
#          It provides comprehensive metadata extraction, analysis, and reporting
#          capabilities for cybersecurity research and penetration testing.
#
# DISCLAIMER: This software is provided "as-is," without warranty of any kind,
#             express or implied, including but not limited to the warranties
#             of merchantability, fitness for a particular purpose, and non-infringement.
#             In no event shall the authors or copyright holders be liable for any claim,
#             damages, or other liability, whether in an action of contract, tort, or otherwise,
#             arising from, out of, or in connection with the software or the use or other dealings
#             in the software.
#
# =============================================================================

import os
import sys
import argparse
import concurrent.futures
import requests
import time
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup
import magic
import logging
from colorama import Fore, Style, init
from rich.console import Console
from rich.table import Table
from pathlib import Path
import re
import ipaddress
from collections import defaultdict
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import json
import subprocess
import shutil
from ipwhois import IPWhois
from datetime import datetime

# Metadata extraction libraries
import PyPDF2
from PIL import Image
from PIL.ExifTags import TAGS
import docx
import openpyxl
import xml.etree.ElementTree as ET
import zipfile
import csv

# Network tools
import dns.resolver
import socket
import whois

# Initialize colorama
init()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("Sidikjari.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger("Sidikjari")
console = Console()

def dns_resolve(domain, record_type):
    """Resolve DNS records with backward compatibility for older dnspython versions"""
    try:
        # Try dnspython 2.x method first
        return dns.resolver.resolve(domain, record_type)
    except AttributeError:
        # Fall back to dnspython 1.x method
        return dns.resolver.query(domain, record_type)

class Sidikjari:
    def __init__(self, target_url=None, output_dir="output", depth=2, threads=10, time_delay=0.0, user_agent="default"):
        # Add https:// scheme if not present and target_url is provided
        if target_url and not target_url.startswith(('http://', 'https://')):
            target_url = f'https://{target_url}'
    
        self.document_metadata = {}  # Stores detailed metadata per document
        self.document_content = {}   # Stores sample content from each document
      
        self.target_url = target_url
        self.output_dir = output_dir
        self.depth = depth
        self.threads = threads
        self.time_delay = time_delay  # Delay between requests in seconds
        self.user_agent = self._get_user_agent(user_agent)  # User agent string
        self.visited_urls = set()
        self.document_urls = set()
        self.file_paths = set()
        
        # Metadata storage
        self.users = set()
        self.emails = set()
        self.software = set()
        self.hosts = set()
        self.internal_domains = set()
        self.ip_addresses = set()
        self.ip_info = {}  # Store detailed IP information
        self.paths = set()
        
        # Initialize exiftool path
        self.exiftool_path = shutil.which('exiftool') or "exiftool"
        
        # File extensions to look for - keep only these specific types
        self.interesting_extensions = {
            'pdf': self.extract_pdf_metadata,
            'docx': self.extract_docx_metadata,
            'xlsx': self.extract_xlsx_metadata,
            'pptx': self.extract_pptx_metadata,
            'jpg': self.extract_image_metadata,
            'jpeg': self.extract_image_metadata,
            'png': self.extract_image_metadata,
            'gif': self.extract_image_metadata,
            'csv': self.extract_csv_metadata
        }
        
        # Create output directory
        os.makedirs(self.output_dir, exist_ok=True)

    def _get_user_agent(self, user_agent_type):
        """Select a user agent based on the specified type"""
        user_agents = {
            "default": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
            "firefox": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0",
            "safari": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.1 Safari/605.1.15",
            "edge": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36 Edg/96.0.1054.62",
            "mobile": "Mozilla/5.0 (iPhone; CPU iPhone OS 15_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.1 Mobile/15E148 Safari/604.1"
        }
        
        if user_agent_type == "random":
            import random
            return random.choice(list(user_agents.values()))
        
        return user_agents.get(user_agent_type, user_agents["default"])

    def crawl_website(self):
        """Crawls the target website to find documents"""
        # Ensure target URL has a scheme
        url = self.target_url
        if url and not url.startswith(('http://', 'https://')):
            url = f'https://{url}'
            self.target_url = url  # Update the target_url with the scheme
            
        logger.info(f"{Fore.GREEN}Starting crawl of {self.target_url}{Style.RESET_ALL}")
        
        self._crawl_url(self.target_url, 0)
        
        logger.info(f"{Fore.GREEN}Crawling complete. Found {len(self.document_urls)} documents{Style.RESET_ALL}")
        
    def _crawl_url(self, url, current_depth):
        """Recursively crawl URLs up to the specified depth, and capture forms"""
        # Ensure URL has a scheme (add https:// if not present)
        if not url.startswith(('http://', 'https://')):
            url = f'https://{url}'
            
        if url in self.visited_urls or current_depth > self.depth:
            return
        
        self.visited_urls.add(url)
        
        try:
            # Implement time delay between requests if specified
            if self.time_delay > 0:
                time.sleep(self.time_delay)
            
            # Set custom headers with the selected user agent
            headers = {
                'User-Agent': self.user_agent
            }
            
            # Disable SSL certificate verification
            response = requests.get(url, timeout=10, verify=False, headers=headers)
            if response.status_code == 200:
                content_type = response.headers.get('Content-Type', '')
                
                # Check if this is a document we're interested in - STRICTLY filter by extension
                file_extension = os.path.splitext(urlparse(url).path)[1].lower().replace('.', '')
                if file_extension in self.interesting_extensions:
                    self.document_urls.add(url)
                    logger.info(f"Found document to analyze: {url} ({file_extension})")
                
                # If HTML, parse links and continue crawling
                if 'text/html' in content_type:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    # Check for forms on the page
                    forms = soup.find_all('form')
                    if forms:
                        logger.info(f"Found {len(forms)} form(s) on {url}")
                        self._capture_form_screenshots(url, forms)
                    
                    for link in soup.find_all('a', href=True):
                        next_url = link['href']
                        
                        # Handle relative URLs
                        if not bool(urlparse(next_url).netloc):
                            next_url = urljoin(url, next_url)
                        
                        # Ensure target_url has a scheme for comparison
                        target_domain = self.target_url
                        if not target_domain.startswith(('http://', 'https://')):
                            target_domain = f'https://{target_domain}'
                        
                        # Only follow links to the same domain
                        if urlparse(target_domain).netloc == urlparse(next_url).netloc:
                            self._crawl_url(next_url, current_depth + 1)
                            
        except Exception as e:
            logger.error(f"Error crawling {url}: {str(e)}")
    
    def _capture_form_screenshots(self, url, forms):
        """Capture screenshots of sensitive forms found on a page using wkhtmltoimage"""
        try:
            # Create a directory for form screenshots if it doesn't exist
            form_screenshots_dir = os.path.join(self.output_dir, "form_screenshots")
            os.makedirs(form_screenshots_dir, exist_ok=True)
            
            # Store form information
            if not hasattr(self, 'form_data'):
                self.form_data = []
            
            # Extract page name from URL for naming screenshots
            parsed_url = urlparse(url)
            page_name = parsed_url.path.strip('/')
            if not page_name:
                page_name = "homepage"
            else:
                # Clean up the page name to be file-system friendly
                page_name = re.sub(r'[^\w\-_]', '_', page_name)
                page_name = re.sub(r'_+', '_', page_name)  # Replace multiple underscores with single
                page_name = page_name[:50]  # Limit length
            
            # Identify sensitive forms
            sensitive_forms = []
            for i, form_element in enumerate(forms):
                # Skip search forms and non-sensitive forms
                if self._is_search_form(form_element):
                    logger.info(f"Skipping search form on {url}")
                    continue
                    
                # Determine if this is a sensitive form
                if self._is_sensitive_form(form_element):
                    # Generate unique identifier for this form
                    form_id = f"{page_name}_{i+1}"
                    
                    # Get form attributes
                    form_attrs = self._extract_form_attributes(form_element)
                    
                    # Create a title for this form
                    form_title = self._create_form_title(form_element, form_attrs, i)
                    
                    # Add form info to our list
                    sensitive_forms.append({
                        'url': url,
                        'form_index': i+1,
                        'form_id': form_id,
                        'title': form_title,
                        'attributes': form_attrs,
                        'page_name': page_name
                    })
            
            # Log how many sensitive forms found
            logger.info(f"Found {len(sensitive_forms)} sensitive forms on {url}")
            
            # Check if wkhtmltoimage is available
            if not shutil.which('wkhtmltoimage'):
                logger.error("wkhtmltoimage tool not found. Please make sure wkhtmltopdf is installed correctly.")
                return
            
            # Only proceed if we have sensitive forms to capture
            if sensitive_forms:
                # Take a screenshot of the full page once
                full_page_screenshot_path = os.path.join(form_screenshots_dir, f"{page_name}_full.png")
                
                # Build the command for full page screenshot
                cmd = [
                    'wkhtmltoimage',
                    '--width', '1366',       # Set width
                    '--height', '1536',      # Set taller height to capture more of page
                    '--quality', '90',       # High quality
                    '--javascript-delay', '2000',  # Wait 2 seconds for JavaScript
                    '--no-stop-slow-scripts',      # Don't stop for slow scripts
                    '--disable-smart-width',       # Use specified width
                    '--enable-local-file-access',  # Allow local file access if needed
                    '--load-error-handling', 'ignore',  # Ignore load errors
                ]
                
                # Add user-agent to avoid bot detection
                cmd.extend(['--custom-header', 'User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'])
                
                # Add the URL and output path
                cmd.extend([url, full_page_screenshot_path])
                
                try:
                    # Execute the command with a timeout
                    process = subprocess.run(
                        cmd,
                        timeout=30,  # 30-second timeout
                        check=False, # Don't raise exception on non-zero exit
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE
                    )
                    
                    # Check if the screenshot was successful
                    if os.path.exists(full_page_screenshot_path) and os.path.getsize(full_page_screenshot_path) > 0:
                        logger.info(f"Full page screenshot saved to {full_page_screenshot_path}")
                        
                        # Use the full page screenshot for all forms
                        for form in sensitive_forms:
                            form['screenshot_path'] = full_page_screenshot_path
                            self.form_data.append(form)
                        
                    else:
                        logger.warning(f"Failed to capture full page screenshot for {url}")
                        
                        # Try with simpler options if the first attempt failed
                        simple_cmd = [
                            'wkhtmltoimage',
                            '--width', '1024',
                            '--height', '1024',
                            '--disable-javascript',  # Disable JavaScript completely
                            url,
                            full_page_screenshot_path
                        ]
                        
                        process = subprocess.run(
                            simple_cmd,
                            timeout=20,
                            check=False,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE
                        )
                        
                        if os.path.exists(full_page_screenshot_path) and os.path.getsize(full_page_screenshot_path) > 0:
                            logger.info(f"Full page screenshot saved using simplified options to {full_page_screenshot_path}")
                            
                            # Add form data with the simplified screenshot
                            for form in sensitive_forms:
                                form['screenshot_path'] = full_page_screenshot_path
                                self.form_data.append(form)
                        else:
                            logger.error(f"Both wkhtmltoimage attempts failed for {url}")
                    
                except subprocess.TimeoutExpired:
                    logger.error(f"Timeout while running wkhtmltoimage for {url}")
                except Exception as wk_e:
                    logger.error(f"Error using wkhtmltoimage: {str(wk_e)}")
                    
        except Exception as e:
            logger.error(f"Error setting up form screenshot capture for {url}: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())
    
    def _is_search_form(self, form_element):
        """Determine if a form is a search form"""
        # Check role attribute
        if form_element.get('role') == 'search':
            return True
            
        # Check class attribute
        if form_element.get('class'):
            classes = form_element['class'] if isinstance(form_element['class'], list) else [form_element['class']]
            if any('search' in c.lower() for c in classes):
                return True
                
        # Check for search input types
        search_inputs = form_element.find_all('input', {'type': 'search'})
        if search_inputs:
            return True
                
        # Check for search in action URL
        action = form_element.get('action', '')
        if 'search' in action.lower():
            return True
                
        # Look for other common search form indicators
        # Fixed the deprecated 'text' argument, replaced with 'string'
        if form_element.find('button', string=re.compile(r'search|find', re.I)):
            return True
            
        return False
    
    def _is_sensitive_form(self, form_element):
        """Determine if a form collects sensitive information"""
        # Login forms typically have password fields
        if form_element.find('input', {'type': 'password'}):
            return True
            
        # Registration/signup forms
        if form_element.find('input', {'name': re.compile(r'register|signup|sign-up|create|account', re.I)}):
            return True
            
        # Contact forms
        contact_fields = form_element.find_all('input', {'name': re.compile(r'name|email|contact|phone|message', re.I)})
        if len(contact_fields) >= 2:
            return True
        
        # Forms with multiple text inputs (likely collecting information)
        text_inputs = form_element.find_all('input', {'type': 'text'})
        if len(text_inputs) >= 3:
            return True
            
        # Forms with textareas (comments, messages, etc.)
        if form_element.find('textarea'):
            return True
            
        # Check for common sensitive form keywords in various attributes
        form_html = str(form_element)
        sensitive_keywords = ['login', 'signin', 'sign-in', 'register', 'signup', 'sign-up', 
                             'contact', 'subscribe', 'newsletter', 'account', 'profile',
                             'checkout', 'payment', 'billing', 'shipping', 'order']
                             
        for keyword in sensitive_keywords:
            if re.search(fr'\b{keyword}\b', form_html, re.I):
                return True
                    
        return False
    
    def _extract_form_attributes(self, form_element):
        """Extract important attributes from a form element"""
        form_attrs = {}
        
        for attr in ['id', 'name', 'action', 'method', 'class', 'role']:
            if form_element.get(attr):
                form_attrs[attr] = form_element[attr]
                
        return form_attrs
    
    def _create_form_title(self, form_element, form_attrs, index):
        """Create a descriptive title for a form"""
        # Start with a default form title
        form_title = f"Form {index+1}"
        
        # Check for common form types
        if form_element.find('input', {'type': 'password'}):
            form_title = "Login Form"
        elif form_element.find('input', {'name': re.compile(r'register|signup|sign-up', re.I)}):
            form_title = "Registration Form"
        elif form_element.find('textarea') and form_element.find('input', {'name': re.compile(r'email', re.I)}):
            form_title = "Contact Form"
        elif form_element.find('input', {'name': re.compile(r'newsletter|subscribe', re.I)}):
            form_title = "Newsletter Subscription"
        elif form_element.find('input', {'name': re.compile(r'checkout|payment|billing', re.I)}):
            form_title = "Payment Form"
            
        # Use ID or name if available and no specific type was identified
        if form_title == f"Form {index+1}":
            if 'id' in form_attrs:
                form_title = f"Form: {form_attrs['id']}"
            elif 'name' in form_attrs:
                form_title = f"Form: {form_attrs['name']}"
        
        return form_title

    def _generate_form_screenshots_section(self, f):
        """Generate a section showing all form screenshots found while crawling"""
        if not hasattr(self, 'form_data') or not self.form_data:
            return
        
        # Create form screenshots section
        f.write("<div class='section'>")
        f.write("<h2>WEBSITE FORMS</h2>")
        
        # Add description
        f.write("<p>The following forms were discovered while crawling the website:</p>")
        
        # Determine how to display forms based on count
        form_count = len(self.form_data)
        
        if form_count <= 3:
            # For few forms, display them full-size
            for form in self.form_data:
                self._generate_single_form_display(f, form)
        else:
            # For many forms, use a thumbnail gallery
            f.write("<div class='gallery' style='display: grid; grid-template-columns: repeat(auto-fill, minmax(250px, 1fr)); gap: 20px;'>")
            
            for form in self.form_data:
                self._generate_form_thumbnail(f, form)
            
            f.write("</div>")
            
            # Add lightbox/modal viewer script
            self._add_lightbox_script(f)
        
        f.write("</div>")  # End of section
    
    def _generate_single_form_display(self, f, form):
        """Generate HTML for displaying a single form in full size"""
        # Get the relative path for HTML embedding
        rel_path = os.path.relpath(form['screenshot_path'], self.output_dir)
        
        f.write(f"<div class='form-container' style='margin-bottom: 30px;'>")
        f.write(f"<h3>{form['title']}</h3>")
        f.write(f"<p>Found on page: <a href='{form['url']}' target='_blank'>{form['url']}</a></p>")
        
        # Display form attributes if available
        if form['attributes']:
            f.write("<p><strong>Form attributes:</strong></p>")
            f.write("<ul>")
            for attr, value in form['attributes'].items():
                f.write(f"<li>{attr}: {value}</li>")
            f.write("</ul>")
        
        # Display the screenshot with link to the form
        f.write(f"""
        <div style="text-align: center; margin: 20px 0;">
            <a href="{form['url']}" target="_blank">
                <img src="{rel_path}" alt="{form['title']}" style="max-width: 100%; border: 1px solid #ddd; border-radius: 5px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);" />
            </a>
        </div>
        """)
        
        f.write("</div>")
    
    def _generate_form_thumbnail(self, f, form):
        """Generate HTML for displaying a form as a thumbnail in a gallery"""
        # Get the relative path for HTML embedding
        rel_path = os.path.relpath(form['screenshot_path'], self.output_dir)
        
        f.write(f"""
        <div class="gallery-item">
            <a href="{form['url']}" target="_blank" class="form-link" data-form-id="{form['form_id']}">
                <div class="thumbnail-container" style="position: relative; overflow: hidden; border-radius: 5px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);">
                    <img src="{rel_path}" alt="{form['title']}" style="width: 100%; height: 180px; object-fit: cover; cursor: pointer;" />
                    <div class="thumbnail-title" style="position: absolute; bottom: 0; left: 0; right: 0; background: rgba(0,0,0,0.7); color: white; padding: 8px; font-size: 14px; text-align: center; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">
                        {form['title']}
                    </div>
                </div>
            </a>
        </div>
        """)
    
    def _add_lightbox_script(self, f):
        """Add JavaScript for a lightbox/modal viewer for the form gallery"""
        f.write("""
        <div id="form-modal" style="display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.8);">
            <div class="modal-content" style="position: relative; margin: 5% auto; padding: 20px; width: 80%; max-width: 1000px; animation: modalopen 0.3s;">
                <span class="close-modal" style="position: absolute; top: 15px; right: 25px; color: white; font-size: 35px; font-weight: bold; cursor: pointer;">&times;</span>
                <div class="modal-body" style="padding: 20px; background: white; border-radius: 5px;">
                    <h3 id="modal-title" style="margin-top: 0;"></h3>
                    <p id="modal-url"></p>
                    <div id="modal-attributes" style="margin-bottom: 15px;"></div>
                    <div style="text-align: center;">
                        <img id="modal-image" style="max-width: 100%; max-height: 70vh; border: 1px solid #ddd;" />
                    </div>
                    <div style="margin-top: 15px; text-align: center;">
                        <a id="modal-link" href="#" target="_blank" class="modal-button" style="display: inline-block; padding: 8px 16px; background-color: #3498db; color: white; text-decoration: none; border-radius: 4px;">
                            Go to Form Page
                        </a>
                    </div>
                </div>
            </div>
        </div>
        
        <script>
        document.addEventListener('DOMContentLoaded', function() {
            // Form data
            const formData = {
        """)
        
        # Add form data as JavaScript object
        for form in self.form_data:
            rel_path = os.path.relpath(form['screenshot_path'], self.output_dir)
            attr_json = json.dumps(form['attributes'])
            
            f.write(f"""
            "{form['form_id']}": {{
                title: "{form['title']}",
                url: "{form['url']}",
                attributes: {attr_json},
                screenshot: "{rel_path}"
            }},
            """)
        
        f.write("""
            };
            
            // Get modal elements
            const modal = document.getElementById('form-modal');
            const modalTitle = document.getElementById('modal-title');
            const modalUrl = document.getElementById('modal-url');
            const modalAttributes = document.getElementById('modal-attributes');
            const modalImage = document.getElementById('modal-image');
            const modalLink = document.getElementById('modal-link');
            const closeModal = document.querySelector('.close-modal');
            
            // Click handler for form links
            document.querySelectorAll('.form-link').forEach(link => {
                link.addEventListener('click', function(e) {
                    e.preventDefault();
                    const formId = this.getAttribute('data-form-id');
                    const form = formData[formId];
                    
                    if (form) {
                        // Populate modal
                        modalTitle.textContent = form.title;
                        modalUrl.textContent = 'Found on: ' + form.url;
                        
                        // Build attributes list
                        let attributesHtml = '';
                        if (Object.keys(form.attributes).length > 0) {
                            attributesHtml = '<strong>Form attributes:</strong><ul>';
                            for (const [key, value] of Object.entries(form.attributes)) {
                                attributesHtml += `<li>${key}: ${value}</li>`;
                            }
                            attributesHtml += '</ul>';
                        }
                        modalAttributes.innerHTML = attributesHtml;
                        
                        // Set image and link
                        modalImage.src = form.screenshot;
                        modalLink.href = form.url;
                        
                        // Show modal
                        modal.style.display = 'block';
                        document.body.style.overflow = 'hidden'; // Prevent scrolling
                    }
                });
            });
            
            // Close modal handlers
            closeModal.addEventListener('click', closeModalFunc);
            window.addEventListener('click', event => {
                if (event.target === modal) {
                    closeModalFunc();
                }
            });
            
            // Close modal function
            function closeModalFunc() {
                modal.style.display = 'none';
                document.body.style.overflow = 'auto'; // Re-enable scrolling
            }
            
            // Close on escape key
            document.addEventListener('keydown', event => {
                if (event.key === 'Escape' && modal.style.display === 'block') {
                    closeModalFunc();
                }
            });
        });
        
        // Add modal animation
        const style = document.createElement('style');
        style.textContent = `
            @keyframes modalopen {
                from {opacity: 0; transform: scale(0.8);}
                to {opacity: 1; transform: scale(1);}
            }
        `;
        document.head.appendChild(style);
        </script>
        """)

    def download_documents(self):
        """Downloads all discovered documents for metadata extraction"""
        logger.info(f"{Fore.GREEN}Downloading {len(self.document_urls)} documents{Style.RESET_ALL}")
        
        documents_dir = os.path.join(self.output_dir, "documents")
        os.makedirs(documents_dir, exist_ok=True)
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
            futures = {executor.submit(self._download_document, doc_url, documents_dir): doc_url for doc_url in self.document_urls}
            
            for future in concurrent.futures.as_completed(futures):
                doc_url = futures[future]
                try:
                    file_path = future.result()
                    if file_path:
                        self.file_paths.add(file_path)
                except Exception as e:
                    logger.error(f"Error downloading {doc_url}: {str(e)}")
        
        logger.info(f"{Fore.GREEN}Downloaded {len(self.file_paths)} documents{Style.RESET_ALL}")
    
    def _download_document(self, url, output_dir):
        """Downloads a single document"""
        try:
            # Implement time delay between requests if specified
            if self.time_delay > 0:
                time.sleep(self.time_delay)
                
            # Set custom headers with the selected user agent
            headers = {
                'User-Agent': self.user_agent
            }
            
            # Disable SSL certificate verification
            response = requests.get(url, timeout=30, stream=True, verify=False, headers=headers)
            if response.status_code == 200:
                # Extract filename from URL
                filename = os.path.basename(urlparse(url).path)
                if not filename:
                    filename = f"document_{hash(url)}"
                
                file_path = os.path.join(output_dir, filename)
                
                with open(file_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                logger.info(f"Downloaded {url} to {file_path}")
                return file_path
            else:
                logger.warning(f"Failed to download {url}, status code: {response.status_code}")
                return None
        except Exception as e:
            logger.error(f"Error downloading {url}: {str(e)}")
            return None

    def extract_all_metadata(self):
        """Extracts metadata from all downloaded files"""
        logger.info(f"{Fore.GREEN}Extracting metadata from {len(self.file_paths)} files{Style.RESET_ALL}")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
            futures = {executor.submit(self._process_file, file_path): file_path for file_path in self.file_paths}
            
            for future in concurrent.futures.as_completed(futures):
                file_path = futures[future]
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {str(e)}")
        
        self._analyze_metadata()
        
    def _process_file(self, file_path):
        """Process a single file to extract metadata"""
        try:
            # Determine file type
            extension = os.path.splitext(file_path)[1].lower().replace('.', '')
            
            # Create document metadata record if it doesn't exist
            if file_path not in self.document_metadata:
                self.document_metadata[file_path] = {
                    'filename': os.path.basename(file_path),
                    'file_path': file_path,
                    'file_size': os.path.getsize(file_path),
                    'file_type': extension,
                    'creation_date': None,
                    'modification_date': None,
                    'authors': set(),
                    'software': set(),
                    'title': None,
                    'subject': None,
                    'keywords': set(),
                    'found_emails': set(),
                    'found_urls': set(),
                    'found_paths': set(),
                    'found_hostnames': set(),
                    'found_ip_addresses': set(),
                    'creation_tool': None,
                    'os_info': None,
                    'gps_data': None,
                    'device_info': None,
                    'all_metadata': {},  # Store ALL metadata fields here
                    'exiftool_metadata': {}  # Store raw exiftool output here
                }
            
            # First run exiftool to get comprehensive metadata
            exif_metadata = self._extract_exiftool_metadata(file_path)
            
            # Get file system metadata
            self._extract_filesystem_metadata(file_path)
                    
            # Then call the specific extractor for additional format-specific extraction
            if extension in self.interesting_extensions:
                self.interesting_extensions[extension](file_path)
            else:
                logger.warning(f"No specific metadata extractor available for {file_path}")
                    
        except Exception as e:
            logger.error(f"Error processing {file_path}: {str(e)}")
    
    def _flatten_metadata(self, metadata, prefix=''):
        """Flatten nested metadata dictionaries for easier access and reporting"""
        result = {}
        if not metadata or not isinstance(metadata, dict):
            return result
            
        for key, value in metadata.items():
            new_key = f"{prefix}{key}" if prefix else key
            if isinstance(value, dict):
                result.update(self._flatten_metadata(value, f"{new_key}."))
            elif isinstance(value, list):
                # Handle lists by converting them to strings
                if all(isinstance(item, dict) for item in value):
                    # If list contains dictionaries, flatten each one
                    for i, item in enumerate(value):
                        result.update(self._flatten_metadata(item, f"{new_key}[{i}]."))
                else:
                    # Otherwise join the list elements
                    result[new_key] = ", ".join(str(item) for item in value)
            else:
                result[new_key] = value
        return result
    
    def _extract_filesystem_metadata(self, file_path):
        """Extract metadata from the file system"""
        try:
            stat_info = os.stat(file_path)
            
            # Creation time (platform dependent)
            if hasattr(stat_info, 'st_birthtime'):  # macOS
                creation_time = datetime.fromtimestamp(stat_info.st_birthtime)
            else:  # Fallback to ctime which might be change time on some systems
                creation_time = datetime.fromtimestamp(stat_info.st_ctime)
            
            # Modification and access times
            modification_time = datetime.fromtimestamp(stat_info.st_mtime)
            access_time = datetime.fromtimestamp(stat_info.st_atime)
            
            # Store in document metadata
            if file_path in self.document_metadata:
                self.document_metadata[file_path]['all_metadata']['filesystem.creation_time'] = creation_time.isoformat()
                self.document_metadata[file_path]['all_metadata']['filesystem.modification_time'] = modification_time.isoformat()
                self.document_metadata[file_path]['all_metadata']['filesystem.access_time'] = access_time.isoformat()
                self.document_metadata[file_path]['all_metadata']['filesystem.size'] = stat_info.st_size
                self.document_metadata[file_path]['all_metadata']['filesystem.permissions'] = stat_info.st_mode
                
                # If no creation date from metadata, use filesystem creation time
                if not self.document_metadata[file_path]['creation_date']:
                    self.document_metadata[file_path]['creation_date'] = creation_time.isoformat()
                    
                # If no modification date from metadata, use filesystem modification time
                if not self.document_metadata[file_path]['modification_date']:
                    self.document_metadata[file_path]['modification_date'] = modification_time.isoformat()
                    
        except Exception as e:
            logger.error(f"Error extracting filesystem metadata for {file_path}: {str(e)}")
    
    def _extract_exiftool_metadata(self, file_path):
        """Extract complete metadata using exiftool"""
        metadata = {}
        try:
            # Run exiftool with all metadata options
            # -a (extract duplicate tags)
            # -u (extract unknown tags)
            # -g (group output by tag category)
            # -j (JSON output)
            # -x (exclude thumbnail data which can be large)
            cmd = [self.exiftool_path, '-a', '-u', '-g', '-j', '-x', 'Thumbnail*', file_path]
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0 and result.stdout:
                # Parse JSON output
                try:
                    exif_data = json.loads(result.stdout)
                    if exif_data and isinstance(exif_data, list) and len(exif_data) > 0:
                        metadata = exif_data[0]
                        
                        # Store all metadata in the document record
                        if file_path in self.document_metadata:
                            # Store the raw exiftool output
                            self.document_metadata[file_path]['exiftool_metadata'] = metadata
                            
                            # Store ALL fields in a flattened structure for easy access
                            flattened = self._flatten_metadata(metadata)
                            self.document_metadata[file_path]['all_metadata'] = flattened
                        
                        # Extract key information for our collections
                        self._process_key_metadata_fields(file_path, metadata)
                        
                        # Log metadata fields found for debugging
                        logger.debug(f"Extracted {len(flattened)} metadata fields from {file_path}")
                        
                except json.JSONDecodeError:
                    logger.error(f"Error parsing exiftool JSON output for {file_path}")
            
            # If we didn't get any metadata, try again with different options
            if not metadata:
                cmd = [self.exiftool_path, '-j', '-a', '-u', file_path]
                result = subprocess.run(cmd, capture_output=True, text=True)
                
                if result.returncode == 0 and result.stdout:
                    try:
                        exif_data = json.loads(result.stdout)
                        if exif_data and isinstance(exif_data, list) and len(exif_data) > 0:
                            metadata = exif_data[0]
                            
                            # Store all metadata in the document record
                            if file_path in self.document_metadata:
                                self.document_metadata[file_path]['exiftool_metadata'] = metadata
                                
                                # Store ALL fields in a flattened structure for easy access
                                flattened = self._flatten_metadata(metadata)
                                self.document_metadata[file_path]['all_metadata'] = flattened
                            
                            # Extract key information for our collections
                            self._process_key_metadata_fields(file_path, metadata)
                    except json.JSONDecodeError:
                        logger.error(f"Error parsing fallback exiftool JSON output for {file_path}")
            
        except Exception as e:
            logger.error(f"Error running exiftool on {file_path}: {str(e)}")
        
        return metadata
    
    def _process_key_metadata_fields(self, file_path, metadata):
        """Process key metadata fields for intelligence gathering"""
        # Flatten nested metadata structure if needed
        flat_metadata = {}
        for group_key, group_data in metadata.items():
            if isinstance(group_data, dict):
                for field_key, field_value in group_data.items():
                    flat_metadata[f"{group_key}:{field_key}"] = field_value
            else:
                flat_metadata[group_key] = group_data
        
        # Look for author/creator information (different naming across file formats)
        author_fields = ['Author', 'Creator', 'Artist', 'Owner', 'By-line', 
                         'OwnerName', 'Microsoft:Author', 'XMP:Creator', 
                         'EXIF:Artist', 'ID3:Artist', 'PDF:Author']
        
        for field in author_fields:
            value = self._get_nested_field(metadata, field)
            if value:
                if isinstance(value, list):
                    for author in value:
                        if author:
                            self.users.add(author)
                            if file_path in self.document_metadata:
                                self.document_metadata[file_path]['authors'].add(author)
                else:
                    self.users.add(value)
                    if file_path in self.document_metadata:
                        self.document_metadata[file_path]['authors'].add(value)
        
        # Look for software information
        software_fields = ['Software', 'Producer', 'CreatorTool', 'Generator', 
                           'Application', 'SourceProgram', 'PDF:Producer', 
                           'XMP:CreatorTool', 'APP14:Adobe']
        
        for field in software_fields:
            value = self._get_nested_field(metadata, field)
            if value:
                if isinstance(value, list):
                    for sw in value:
                        if sw:
                            self.software.add(sw)
                            if file_path in self.document_metadata:
                                self.document_metadata[file_path]['software'].add(sw)
                else:
                    self.software.add(value)
                    if file_path in self.document_metadata:
                        self.document_metadata[file_path]['software'].add(value)
        
        # Look for title information
        title_fields = ['Title', 'DocumentName', 'Headline', 'ObjectName', 
                        'XMP:Title', 'PDF:Title', 'ID3:Title']
        
        for field in title_fields:
            value = self._get_nested_field(metadata, field)
            if value and file_path in self.document_metadata:
                self.document_metadata[file_path]['title'] = value
                break
        
        # Look for subject/description information
        subject_fields = ['Subject', 'Description', 'Caption', 'Comment', 
                          'XMP:Description', 'PDF:Subject', 'ID3:Comment']
        
        for field in subject_fields:
            value = self._get_nested_field(metadata, field)
            if value and file_path in self.document_metadata:
                self.document_metadata[file_path]['subject'] = value
                break
        
        # Look for dates
        creation_date_fields = ['CreateDate', 'DateTimeOriginal', 'CreationDate', 
                               'DateCreated', 'PDF:CreationDate', 'XMP:CreateDate']
        
        for field in creation_date_fields:
            value = self._get_nested_field(metadata, field)
            if value and file_path in self.document_metadata:
                self.document_metadata[file_path]['creation_date'] = value
                break
        
        modification_date_fields = ['ModifyDate', 'FileModifyDate', 'ModificationDate', 
                                   'PDF:ModDate', 'XMP:ModifyDate']
        
        for field in modification_date_fields:
            value = self._get_nested_field(metadata, field)
            if value and file_path in self.document_metadata:
                self.document_metadata[file_path]['modification_date'] = value
                break
        
        # Extract GPS coordinates if available
        gps_fields = {
            'lat': ['GPSLatitude', 'GPS:GPSLatitude', 'XMP:GPSLatitude'],
            'lon': ['GPSLongitude', 'GPS:GPSLongitude', 'XMP:GPSLongitude'],
            'alt': ['GPSAltitude', 'GPS:GPSAltitude', 'XMP:GPSAltitude']
        }
        
        gps_data = {}
        for coord_type, fields in gps_fields.items():
            for field in fields:
                value = self._get_nested_field(metadata, field)
                if value:
                    gps_data[coord_type] = value
                    break
        
        if gps_data and file_path in self.document_metadata:
            self.document_metadata[file_path]['gps_data'] = gps_data
        
        # Look for device information
        device_fields = ['Model', 'Make', 'DeviceManufacturer', 'DeviceModel', 
                         'EXIF:Make', 'EXIF:Model', 'XMP:Device']
        
        device_info = {}
        for field in device_fields:
            value = self._get_nested_field(metadata, field)
            if value:
                device_info[field] = value
        
        if device_info and file_path in self.document_metadata:
            self.document_metadata[file_path]['device_info'] = device_info
    
    def _get_nested_field(self, metadata, field_path):
        """Get a field value from potentially nested metadata structure"""
        parts = field_path.split(':')
        
        # Direct field access
        if len(parts) == 1 and field_path in metadata:
            return metadata[field_path]
        
        # Nested field access
        if len(parts) > 1:
            current = metadata
            for part in parts:
                if part in current:
                    current = current[part]
                else:
                    return None
            return current
        
        return None

    def extract_csv_metadata(self, file_path):
        """Extract metadata from CSV files"""
        try:
            with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
                csv_reader = csv.reader(f)
                for row in csv_reader:
                    for cell in row:
                        if cell and isinstance(cell, str):
                            self._extract_from_text(cell)
        
        except Exception as e:
            logger.error(f"Error extracting CSV metadata from {file_path}: {str(e)}")

    def extract_pdf_metadata(self, file_path):
        """Extract enhanced metadata from PDF files"""
        try:
            with open(file_path, 'rb') as f:
                pdf = PyPDF2.PdfReader(f)
                info = pdf.metadata
                
                # Create document metadata record if it doesn't exist
                if file_path not in self.document_metadata:
                    self.document_metadata[file_path] = {
                        'filename': os.path.basename(file_path),
                        'file_path': file_path,
                        'file_size': os.path.getsize(file_path),
                        'file_type': 'pdf',
                        'creation_date': None,
                        'modification_date': None,
                        'authors': set(),
                        'software': set(),
                        'title': None,
                        'subject': None,
                        'keywords': set(),
                        'found_emails': set(),
                        'found_urls': set(),
                        'found_paths': set(),
                        'found_hostnames': set(),
                        'found_ip_addresses': set(),
                        'all_metadata': {},  # Store ALL metadata fields here
                        'exiftool_metadata': {}  # Store raw exiftool output here
                    }
                
                # Process PDF metadata if available
                if info:
                    # Basic metadata fields
                    if hasattr(info, 'author') and info.author:
                        self.document_metadata[file_path]['authors'].add(info.author)
                        self.users.add(info.author)
                    
                    if hasattr(info, 'creator') and info.creator:
                        self.document_metadata[file_path]['software'].add(info.creator)
                        self.software.add(info.creator)
                        # Look for potential usernames in creator field
                        if '\\' in info.creator:
                            parts = info.creator.split('\\')
                            if len(parts) > 1:
                                self.users.add(parts[1])
                                self.document_metadata[file_path]['authors'].add(parts[1])
                    
                    if hasattr(info, 'title') and info.title:
                        self.document_metadata[file_path]['title'] = info.title
                    
                    if hasattr(info, 'subject') and info.subject:
                        self.document_metadata[file_path]['subject'] = info.subject
                    
                    if hasattr(info, 'producer') and info.producer:
                        self.document_metadata[file_path]['software'].add(info.producer)
                        self.software.add(info.producer)
                    
                    # Parse creation and modification dates
                    if '/CreationDate' in info:
                        date_str = info['/CreationDate']
                        if isinstance(date_str, str) and date_str.startswith('D:'):
                            self.document_metadata[file_path]['creation_date'] = date_str[2:14]  # Extract date part
                    
                    if '/ModDate' in info:
                        date_str = info['/ModDate']
                        if isinstance(date_str, str) and date_str.startswith('D:'):
                            self.document_metadata[file_path]['modification_date'] = date_str[2:14]  # Extract date part
                    
                    # Check for all metadata fields - process dictionary
                    for key in info:
                        try:
                            value = info[key]
                            if isinstance(value, str):
                                # Store in all_metadata
                                self.document_metadata[file_path]['all_metadata'][key] = value
                                
                                # Extract emails, URLs, and paths from metadata
                                emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', value)
                                urls = re.findall(r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+', value)
                                paths = re.findall(r'[A-Za-z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*[^\\/:*?"<>|\r\n]*', value)
                                
                                if emails:
                                    self.document_metadata[file_path]['found_emails'].update(emails)
                                    self.emails.update(emails)
                                
                                if urls:
                                    self.document_metadata[file_path]['found_urls'].update(urls)
                                
                                if paths:
                                    self.document_metadata[file_path]['found_paths'].update(paths)
                                    self.paths.update(paths)
                        except Exception as sub_e:
                            logger.debug(f"Error processing metadata field {key}: {sub_e}")
                
                # Also run exiftool for more comprehensive metadata
                exiftool_metadata = self._extract_exiftool_metadata(file_path)
                if exiftool_metadata:
                    # Merge with existing metadata
                    if 'all_metadata' not in self.document_metadata[file_path]:
                        self.document_metadata[file_path]['all_metadata'] = {}
                    
                    flat_metadata = self._flatten_metadata(exiftool_metadata)
                    self.document_metadata[file_path]['all_metadata'].update(flat_metadata)
                
                # Extract text content for further analysis
                extracted_text = ""
                try:
                    for page_num in range(len(pdf.pages)):
                        page = pdf.pages[page_num]
                        text = page.extract_text()
                        if text:
                            extracted_text += text + "\n"
                            self._extract_from_text(text)
                except Exception as text_e:
                    logger.debug(f"Error extracting text from PDF {file_path}: {text_e}")
                
                # Store summary of extracted text for later analysis
                text_sample = extracted_text[:2000] if len(extracted_text) > 2000 else extracted_text
                self.document_content[file_path] = text_sample
                    
        except Exception as e:
            logger.error(f"Error extracting PDF metadata from {file_path}: {str(e)}")

    def extract_docx_metadata(self, file_path):
        """Extract metadata from DOCX files"""
        try:
            doc = docx.Document(file_path)
            core_props = doc.core_properties
            
            # Extract creator info
            if core_props.author:
                self.users.add(core_props.author)
            if core_props.last_modified_by:
                self.users.add(core_props.last_modified_by)
                
            # Extract app version
            if hasattr(core_props, 'revision') and core_props.revision:
                self.software.add(f"Microsoft Word - {core_props.revision}")
            
            # Extract text for further analysis
            for para in doc.paragraphs:
                self._extract_from_text(para.text)
            
        except Exception as e:
            logger.error(f"Error extracting DOCX metadata from {file_path}: {str(e)}")

    def extract_xlsx_metadata(self, file_path):
        """Extract metadata from XLSX files"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Extract metadata from workbook properties
            if wb.properties.creator:
                self.users.add(wb.properties.creator)
            if wb.properties.lastModifiedBy:
                self.users.add(wb.properties.lastModifiedBy)
            
            # Extract text from each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            self._extract_from_text(cell.value)
            
        except Exception as e:
            logger.error(f"Error extracting XLSX metadata from {file_path}: {str(e)}")

    def extract_pptx_metadata(self, file_path):
        """Extract metadata from PPTX files"""
        try:
            # PPTX files are ZIP files with XML content
            with zipfile.ZipFile(file_path) as zf:
                # Extract core properties
                if 'docProps/core.xml' in zf.namelist():
                    with zf.open('docProps/core.xml') as f:
                        xml_content = f.read()
                        root = ET.fromstring(xml_content)
                        
                        # Extract creator
                        creator = root.find('.//{http://purl.org/dc/elements/1.1/}creator')
                        if creator is not None and creator.text:
                            self.users.add(creator.text)
                        
                        # Extract last modified by
                        last_modified_by = root.find('.//{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}lastModifiedBy')
                        if last_modified_by is not None and last_modified_by.text:
                            self.users.add(last_modified_by.text)
                
                # Extract app properties
                if 'docProps/app.xml' in zf.namelist():
                    with zf.open('docProps/app.xml') as f:
                        xml_content = f.read()
                        root = ET.fromstring(xml_content)
                        
                        # Extract application
                        application = root.find('.//{http://schemas.openxmlformats.org/officeDocument/2006/extended-properties}Application')
                        if application is not None and application.text:
                            self.software.add(application.text)
                
                # Extract slide content
                for name in zf.namelist():
                    if re.match(r'ppt/slides/slide[0-9]+\.xml', name):
                        with zf.open(name) as f:
                            xml_content = f.read()
                            root = ET.fromstring(xml_content)
                            
                            # Extract text from each text run in slide
                            for text_node in root.findall('.//*[@type="txBody"]//a:t', 
                                                        namespaces={'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}):
                                if text_node.text:
                                    self._extract_from_text(text_node.text)
        
        except Exception as e:
            logger.error(f"Error extracting PPTX metadata from {file_path}: {str(e)}")

    def extract_image_metadata(self, file_path):
        """Extract metadata from image files (EXIF data)"""
        try:
            with Image.open(file_path) as img:
                exif_data = img._getexif()
                
                if exif_data:
                    for tag_id, value in exif_data.items():
                        tag = TAGS.get(tag_id, tag_id)
                        
                        # Look for interesting EXIF tags
                        if tag in ['Make', 'Model', 'Software']:
                            self.software.add(f"{tag}: {value}")
                        elif tag in ['Artist', 'Copyright', 'ImageDescription']:
                            if isinstance(value, str):
                                self.users.add(value)
                                self._extract_from_text(value)
                        
                        # Some cameras/phones store GPS data
                        elif tag == 'GPSInfo':
                            # Process GPS data if needed
                            pass
        
        except Exception as e:
            logger.error(f"Error extracting image metadata from {file_path}: {str(e)}")

    def _extract_from_text(self, text):
        """Extract useful information from text content"""
        if not text or not isinstance(text, str):
            return
        
        # Extract email addresses
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, text)
        for email in emails:
            self.emails.add(email)
            # Extract domain from email
            domain = email.split('@')[1]
            self.internal_domains.add(domain)
        
        # Extract potential internal domain names
        domain_pattern = r'(?:https?://)?(?:www\.)?([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})(?:/[^\s]*)?'
        domains = re.findall(domain_pattern, text)
        for domain in domains:
            if not any(public_domain in domain for public_domain in ['google.com', 'microsoft.com', 'yahoo.com']):
                self.internal_domains.add(domain)
        
        # Extract potential file paths
        path_pattern = r'[A-Za-z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*[^\\/:*?"<>|\r\n]*'
        paths = re.findall(path_pattern, text)
        for path in paths:
            self.paths.add(path)
            # Extract potential username from path
            if 'Users\\' in path:
                parts = path.split('Users\\')
                if len(parts) > 1:
                    user_path = parts[1].split('\\')[0]
                    if user_path and user_path not in ['Public', 'All Users', 'Default']:
                        self.users.add(user_path)
        
        # Extract potential IP addresses
        ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
        ips = re.findall(ip_pattern, text)
        for ip in ips:
            try:
                # Validate IP address
                ipaddress.ip_address(ip)
                # Exclude common non-internal IPs
                if not ip.startswith(('127.', '255.', '0.')):
                    self.ip_addresses.add(ip)
            except ValueError:
                pass
        
        # Extract hostnames (server names)
        hostname_pattern = r'\b([a-zA-Z0-9-]{2,}(?:\.[a-zA-Z0-9-]+)*)\b'
        for match in re.finditer(hostname_pattern, text):
            hostname = match.group(1)
            if len(hostname) > 2 and not any(c.isdigit() for c in hostname):
                if re.match(r'^[a-zA-Z0-9-]+$', hostname):
                    # Exclude common words
                    common_words = ['http', 'https', 'www', 'com', 'net', 'org']
                    if hostname.lower() not in common_words:
                        self.hosts.add(hostname)

    def _analyze_metadata(self):
        """Analyze collected metadata to find relationships"""
        logger.info(f"{Fore.GREEN}Analyzing collected metadata{Style.RESET_ALL}")

        # Group data by domain
        domain_data = defaultdict(lambda: {
            'users': set(),
            'emails': set(),
            'hosts': set(),
            'ips': set(),
            'software': set()
        })

        # Process emails
        for email in self.emails:
            if '@' in email:
                username, domain = email.split('@')
                domain_data[domain]['users'].add(username)
                domain_data[domain]['emails'].add(email)

        # Process domains and IPs
        for domain in self.internal_domains:
            try:
                # Try to resolve domain to IP
                answers = dns_resolve(domain, 'A')
                for answer in answers:
                    ip = answer.to_text()
                    domain_data[domain]['ips'].add(ip)
                    self.ip_addresses.add(ip)
            except:
                pass

        # Generate domain report
        self.generate_reports(domain_data)

    def generate_reports(self, domain_data=None):
        """Generate HTML report with enhanced features"""
        logger.info(f"{Fore.GREEN}Generating report{Style.RESET_ALL}")
        
        # Store domain_data as a class attribute if provided
        if domain_data:
            self.domain_data = domain_data
        
        # Determine the target domain from target_url if available
        target_domain = None
        if self.target_url:
            parsed_url = urlparse(self.target_url)
            target_domain = parsed_url.netloc.lower()
            if target_domain.startswith('www.'):
                target_domain = target_domain[4:]
        
        # If we have a target domain, preprocess WHOIS data
        domain_info = None
        if target_domain:
            try:
                # Force WHOIS data collection before report generation
                domain_info = self._analyze_domain_info(target_domain)
                logger.info(f"Collected WHOIS data for {target_domain}")
            except Exception as e:
                logger.error(f"Error collecting WHOIS data: {str(e)}")
        
        # Generate HTML report
        report_filename = "Sidikjari_report.html"
        report_path = os.path.join(self.output_dir, report_filename)
        
        try:
            self._generate_html_report(report_path, target_domain, domain_info)
            logger.info(f"{Fore.GREEN}Report generated: {report_path}{Style.RESET_ALL}")
            return [report_path]
        except Exception as e:
            logger.error(f"Error generating HTML report: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())
            return []

    def _analyze_domain_info(self, domain):
        """Gather comprehensive information about a specific domain"""
        # Initialize structure with all fields set to None or empty lists
        domain_info = {
            'registrant': {
                'name': None,
                'organization': None,
                'email': None,
                'phone': None,
                'fax': None,
                'street': None,
                'city': None,
                'state': None,
                'postal_code': None,
                'country': None
            },
            'admin': {
                'name': None,
                'organization': None,
                'email': None,
                'phone': None,
                'fax': None,
                'street': None,
                'city': None,
                'state': None,
                'postal_code': None,
                'country': None
            },
            'tech': {
                'name': None,
                'organization': None,
                'email': None,
                'phone': None,
                'fax': None,
                'street': None,
                'city': None,
                'state': None,
                'postal_code': None,
                'country': None
            },
            'registrar': None,
            'creation_date': None,
            'update_date': None,
            'expiration_date': None,
            'name_servers': [],
            'domain_status': [],
            'ip_addresses': [],
            'mx_records': []
        }

        # Get WHOIS information
        try:
            logger.info(f"Getting WHOIS information for {domain}")
            w = whois.whois(domain)
            logger.debug(f"Raw WHOIS data: {w}")

            # Convert data to a dictionary for easier handling
            whois_dict = {}
            for key, value in w.items():
                if value is not None:
                    whois_dict[key.lower()] = value

            # Process standard fields
            if 'registrar' in whois_dict:
                domain_info['registrar'] = whois_dict['registrar']

            # Process dates
            for date_field, target_field in [
                ('creation_date', 'creation_date'),
                ('updated_date', 'update_date'),
                ('expiration_date', 'expiration_date')
            ]:
                if date_field in whois_dict:
                    value = whois_dict[date_field]
                    if isinstance(value, list) and len(value) > 0:
                        domain_info[target_field] = value[0]
                    else:
                        domain_info[target_field] = value

            # Process name servers
            if 'name_servers' in whois_dict:
                ns_list = whois_dict['name_servers']
                if isinstance(ns_list, list):
                    domain_info['name_servers'] = ns_list
                elif ns_list:
                    domain_info['name_servers'] = [ns_list]

            # Process domain status
            if 'status' in whois_dict:
                status_list = whois_dict['status']
                if isinstance(status_list, list):
                    domain_info['domain_status'] = status_list
                elif status_list:
                    domain_info['domain_status'] = [status_list]

            # Process contact information directly from dict keys
            contact_mappings = {
                'registrant': {
                    'name': ['registrant_name', 'registrant'],
                    'organization': ['registrant_organization', 'org', 'organization'],
                    'email': ['registrant_email', 'email'],
                    'phone': ['registrant_phone', 'phone'],
                    'fax': ['registrant_fax', 'fax'],
                    'street': ['registrant_street', 'address', 'street'],
                    'city': ['registrant_city', 'city'],
                    'state': ['registrant_state_province', 'state', 'province'],
                    'postal_code': ['registrant_postal_code', 'postal_code', 'zip'],
                    'country': ['registrant_country', 'country']
                },
                'admin': {
                    'name': ['admin_name', 'administrative_name'],
                    'organization': ['admin_organization', 'administrative_organization'],
                    'email': ['admin_email', 'administrative_email'],
                    'phone': ['admin_phone', 'administrative_phone'],
                    'fax': ['admin_fax', 'administrative_fax'],
                    'street': ['admin_street', 'administrative_street', 'administrative_address'],
                    'city': ['admin_city', 'administrative_city'],
                    'state': ['admin_state_province', 'administrative_state_province', 'administrative_state'],
                    'postal_code': ['admin_postal_code', 'administrative_postal_code', 'administrative_zip'],
                    'country': ['admin_country', 'administrative_country']
                },
                'tech': {
                    'name': ['tech_name', 'technical_name'],
                    'organization': ['tech_organization', 'technical_organization'],
                    'email': ['tech_email', 'technical_email'],
                    'phone': ['tech_phone', 'technical_phone'],
                    'fax': ['tech_fax', 'technical_fax'],
                    'street': ['tech_street', 'technical_street', 'technical_address'],
                    'city': ['tech_city', 'technical_city'],
                    'state': ['tech_state_province', 'technical_state_province', 'technical_state'],
                    'postal_code': ['tech_postal_code', 'technical_postal_code', 'technical_zip'],
                    'country': ['tech_country', 'technical_country']
                }
            }

            # Process contact information with improved mapping
            for contact_type, fields in contact_mappings.items():
                for field, key_options in fields.items():
                    for key in key_options:
                        if key.lower() in whois_dict:
                            domain_info[contact_type][field] = whois_dict[key.lower()]
                            break

            # Process raw text data using regex if available
            if hasattr(w, 'text') and w.text:
                whois_text = w.text.lower()
                logger.debug(f"Processing raw WHOIS text: {whois_text[:200]}...")  # Log first 200 chars

                # Process contact information using regex for each contact type
                for contact_type in ['registrant', 'admin', 'tech']:
                    for field in domain_info[contact_type].keys():
                        # Only try to extract data if it's not already set
                        if not domain_info[contact_type][field]:
                            # Try different pattern formats
                            patterns = [
                                rf"{contact_type}\s+{field}:\s*([^\n]+)",
                                rf"{contact_type} {field}:\s*([^\n]+)",
                                rf"{contact_type}-{field}:\s*([^\n]+)"
                            ]

                            for pattern in patterns:
                                match = re.search(pattern, whois_text)
                                if match:
                                    domain_info[contact_type][field] = match.group(1).strip()
                                    break

                # Try to extract registrar info if not already set
                if not domain_info['registrar']:
                    registrar_match = re.search(r"registrar:\s*([^\n]+)", whois_text)
                    if registrar_match:
                        domain_info['registrar'] = registrar_match.group(1).strip()

                # Try to extract name servers if not already set
                if not domain_info['name_servers']:
                    ns_matches = re.findall(r"name server:\s*([^\n]+)", whois_text)
                    if ns_matches:
                        domain_info['name_servers'] = [ns.strip() for ns in ns_matches]
        except Exception as e:
            logger.error(f"Error getting WHOIS information for {domain}: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())

        # Get DNS A records
        try:
            logger.info(f"Getting DNS A records for {domain}")
            # First try to get all A records
            answers = dns_resolve(domain, 'A')
            for answer in answers:
                ip = answer.to_text()
                domain_info['ip_addresses'].append(ip)

                # Get additional IP information
                self._get_ip_info(ip, domain)

            # Also check www. subdomain
            try:
                www_domain = f"www.{domain}"
                www_answers = dns_resolve(www_domain, 'A')
                for answer in www_answers:
                    ip = answer.to_text()
                    if ip not in domain_info['ip_addresses']:
                        domain_info['ip_addresses'].append(ip)
                        self._get_ip_info(ip, domain)
            except Exception as www_e:
                logger.debug(f"Error resolving www.{domain}: {str(www_e)}")
        except Exception as dns_e:
            logger.error(f"Error resolving DNS A records for {domain}: {str(dns_e)}")

        # Get MX records
        try:
            logger.info(f"Getting DNS MX records for {domain}")
            mx_records = dns_resolve(domain, 'MX')
            for mx in mx_records:
                domain_info['mx_records'].append(f"{mx.preference} {mx.exchange}")
        except Exception as mx_e:
            logger.debug(f"Error resolving MX records for {domain}: {str(mx_e)}")

        # Log the collected information
        logger.info(f"Completed domain info collection for {domain}")

        return domain_info

    def _map_contact_attribute(self, contact_dict, attr_name, attr_value):
        """Map WHOIS attributes to contact fields"""
        attr_lower = attr_name.lower()
        
        # Process name fields
        if any(name_field in attr_lower for name_field in ['name', 'registrant', 'admin', 'tech']):
            if 'name' in attr_lower or attr_lower in ['registrant', 'admin', 'tech']:
                contact_dict['name'] = attr_value
        
        # Process organization fields
        elif any(org_field in attr_lower for org_field in ['organization', 'org', 'company']):
            contact_dict['organization'] = attr_value
        
        # Process email fields
        elif 'email' in attr_lower:
            contact_dict['email'] = attr_value
        
        # Process phone fields
        elif 'phone' in attr_lower:
            contact_dict['phone'] = attr_value
        
        # Process address fields
        elif any(addr_field in attr_lower for addr_field in ['address', 'street', 'city', 'state', 'country']):
            if contact_dict['address']:
                contact_dict['address'] = f"{contact_dict['address']}, {attr_value}"
            else:
                contact_dict['address'] = attr_value

    def _get_ip_info(self, ip, associated_domain=None):
        """Get detailed information about an IP address"""
        if ip in self.ip_info:
            return self.ip_info[ip]
            
        ip_data = {
            'cidr': None,
            'asn': None,
            'organization': None,
            'country': None,
            'associated_domains': set(),
            'reverse_dns': None
        }
        
        if associated_domain:
            ip_data['associated_domains'].add(associated_domain)
        
        # Try to get WHOIS information for the IP
        try:
            ip_whois = IPWhois(ip)
            results = ip_whois.lookup_rdap()
            
            if results:
                if 'network' in results and 'cidr' in results['network']:
                    ip_data['cidr'] = results['network']['cidr']
                
                if 'asn' in results:
                    ip_data['asn'] = results['asn']
                    
                if 'asn_description' in results:
                    ip_data['organization'] = results['asn_description']
                    
                if 'network' in results and 'country' in results['network']:
                    ip_data['country'] = results['network']['country']
        except Exception as e:
            logger.warning(f"Error getting WHOIS information for IP {ip}: {str(e)}")
        
        # Try reverse DNS lookup
        try:
            hostname, _, _ = socket.gethostbyaddr(ip)
            ip_data['reverse_dns'] = hostname
        except:
            pass
        
        self.ip_info[ip] = ip_data
        return ip_data

    def _generate_html_report(self, report_path, target_domain, domain_info=None):

        try:
            # Group documents by file type
            documents_by_type = {}
            for file_path, metadata in self.document_metadata.items():
                file_type = metadata['file_type']
                if file_type not in documents_by_type:
                    documents_by_type[file_type] = []
                documents_by_type[file_type].append((file_path, metadata))
            
            with open(report_path, 'w') as f:
                # HTML header with background image
                f.write("""<!DOCTYPE html>
    <html>
    <head>
        <title>Sidikjari Metadata Analysis Report</title>
        <style>
            body { 
                font-family: Arial, sans-serif; 
                margin: 20px; 
                background-image: url('https://static.wixstatic.com/media/488c5b_8bd517d20d2b446e906385dec6bf1898~mv2.jpg');
                background-attachment: fixed;
                background-size: contain;
                background-repeat: no-repeat;
                background-position: center;
                background-color: #ffffff;
                position: relative;
            }
            body::before {
                content: "";
                position: fixed;
                top: 0;
                left: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(255, 255, 255, 0.8); /* 80% opacity white overlay */
                z-index: -1;
            }
            h1 { color: #2c3e50; }
            h2 { color: #3498db; margin-top: 30px; }
            h3 { color: #2980b9; }
            .container { 
                max-width: 1200px; 
                margin: 0 auto; 
                background-color: rgba(255, 255, 255, 0.9);
                padding: 20px;
                border-radius: 10px;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
            }
            .section { margin-bottom: 30px; padding: 20px; border: 1px solid #ddd; border-radius: 5px; background-color: #ffffff; }
            .metadata-item { margin-bottom: 20px; padding: 10px; background-color: #f9f9f9; border-radius: 5px; }
            table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
            th, td { padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }
            th { background-color: #f2f2f2; }
            .footer { margin-top: 50px; text-align: center; font-size: 12px; color: #7f8c8d; }
            .metadata-table { font-size: 12px; }
            .key-column { width: 40%; font-weight: bold; }
            .value-column { width: 60%; word-break: break-word; }
            .debug-info { margin: 20px; padding: 10px; background-color: #ffe0e0; border: 1px solid #ffcccc; display: none; }
            
            /* Collapsible section styles */
            .collapsible {
                background-color: #3498db;
                color: white;
                cursor: pointer;
                padding: 12px;
                width: 100%;
                border: none;
                text-align: left;
                outline: none;
                font-size: 16px;
                font-weight: bold;
                border-radius: 5px 5px 0 0;
                margin-top: 20px;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            .active, .collapsible:hover {
                background-color: #2980b9;
            }
            .document-type-content {
                padding: 0 18px;
                max-height: 0;
                overflow: hidden;
                transition: max-height 0.3s ease-out;
                background-color: #f9f9f9;
                border-radius: 0 0 5px 5px;
                border: 1px solid #ddd;
                border-top: none;
            }
            .document-count {
                background-color: white;
                color: #3498db;
                border-radius: 50%;
                padding: 2px 8px;
                font-size: 14px;
            }
            /* CSS for toggle icon */
            .collapsible:after {
                content: '\\02795'; /* Unicode character for "plus" sign (+) */
                font-size: 13px;
                color: white;
                margin-left: 5px;
            }
            .active:after {
                content: "\\2796"; /* Unicode character for "minus" sign (-) */
            }
        </style>
        <script>
            document.addEventListener('DOMContentLoaded', function() {
                var coll = document.getElementsByClassName("collapsible");
                for (var i = 0; i < coll.length; i++) {
                    coll[i].addEventListener("click", function() {
                        this.classList.toggle("active");
                        var content = this.nextElementSibling;
                        if (content.style.maxHeight) {
                            content.style.maxHeight = null;
                        } else {
                            content.style.maxHeight = content.scrollHeight + "px";
                        }
                    });
                }
                
                // Expand the first section by default
                if (coll.length > 0) {
                    coll[0].click();
                }
            });
        </script>
    </head>
    <body>
        <div class="container">
            <h1>Sidikjari Metadata Analysis Report</h1>
    """)
    
                # Target information
                f.write(f"<p><strong>Target:</strong> {self.target_url if self.target_url else self.input_dir}</p>")
                
                # Website Screenshot section (if applicable)
                if self.target_url:
                    self._generate_screenshot_section(f, self.target_url)
                    
                    # Add form screenshots section if forms were found
                    if hasattr(self, 'form_data') and self.form_data:
                        self._generate_form_screenshots_section(f)
                
                # Domain Information
                if target_domain:
                    # If domain_info was not provided, try to get it now
                    if domain_info is None:
                        try:
                            domain_info = self._analyze_domain_info(target_domain)
                        except Exception as e:
                            f.write(f"<div class='debug-info'>Error collecting domain info: {str(e)}</div>")
                    
                    if domain_info:
                        f.write("<div class='section'>")
                        f.write("<h2>DOMAIN INFORMATION</h2>")
                        f.write(f"<p><strong>Domain:</strong> {target_domain}</p>")
                        
                        # Registrant Information
                        f.write("<h3>Registrant Information</h3>")
                        
                        # Debug info - uncomment by changing display:none to display:block in CSS
                        f.write("<div class='debug-info'>")
                        f.write("<strong>Debug:</strong> Registrant fields available: ")
                        f.write(", ".join([k for k, v in domain_info['registrant'].items() if v]))
                        f.write("</div>")
                        
                        f.write("<table>")
                        
                        for field, label in [
                            ('name', 'Name'),
                            ('organization', 'Organization'),
                            ('email', 'Email'),
                            ('phone', 'Phone'),
                            ('fax', 'Fax'),
                            ('street', 'Street'),
                            ('city', 'City'),
                            ('state', 'State/Province'),
                            ('postal_code', 'Postal Code'),
                            ('country', 'Country')
                        ]:
                            value = domain_info['registrant'].get(field)
                            if value:
                                f.write(f"<tr><td>{label}</td><td>{value}</td></tr>")
                        
                        # If no registrant data was found, display a message
                        if not any(domain_info['registrant'].values()):
                            f.write("<tr><td colspan='2'>No registrant information available</td></tr>")
                        
                        f.write("</table>")
                        
                        # Admin Information
                        f.write("<h3>Admin Information</h3>")
                        f.write("<table>")
                        
                        for field, label in [
                            ('name', 'Name'),
                            ('organization', 'Organization'),
                            ('email', 'Email'),
                            ('phone', 'Phone'),
                            ('fax', 'Fax'),
                            ('street', 'Street'),
                            ('city', 'City'),
                            ('state', 'State/Province'),
                            ('postal_code', 'Postal Code'),
                            ('country', 'Country')
                        ]:
                            value = domain_info['admin'].get(field)
                            if value:
                                f.write(f"<tr><td>{label}</td><td>{value}</td></tr>")
                        
                        # If no admin data was found, display a message
                        if not any(domain_info['admin'].values()):
                            f.write("<tr><td colspan='2'>No admin information available</td></tr>")
                        
                        f.write("</table>")
                        
                        # Tech Information
                        f.write("<h3>Tech Information</h3>")
                        f.write("<table>")
                        
                        for field, label in [
                            ('name', 'Name'),
                            ('organization', 'Organization'),
                            ('email', 'Email'),
                            ('phone', 'Phone'),
                            ('fax', 'Fax'),
                            ('street', 'Street'),
                            ('city', 'City'),
                            ('state', 'State/Province'),
                            ('postal_code', 'Postal Code'),
                            ('country', 'Country')
                        ]:
                            value = domain_info['tech'].get(field)
                            if value:
                                f.write(f"<tr><td>{label}</td><td>{value}</td></tr>")
                        
                        # If no tech data was found, display a message
                        if not any(domain_info['tech'].values()):
                            f.write("<tr><td colspan='2'>No tech information available</td></tr>")
                        
                        f.write("</table>")
                        
                        # General domain information
                        f.write("<h3>Domain Details</h3>")
                        
                        # Debug info - uncomment by changing display:none to display:block in CSS
                        f.write("<div class='debug-info'>")
                        f.write("<strong>Debug:</strong> Domain fields available: ")
                        f.write(", ".join([k for k, v in domain_info.items() if v and not isinstance(v, dict)]))
                        f.write("</div>")
                        
                        f.write("<table>")
                        
                        if domain_info.get('registrar'):
                            f.write(f"<tr><td>Registrar</td><td>{domain_info['registrar']}</td></tr>")
                        
                        if domain_info.get('creation_date'):
                            f.write(f"<tr><td>Creation Date</td><td>{domain_info['creation_date']}</td></tr>")
                        
                        if domain_info.get('update_date'):
                            f.write(f"<tr><td>Updated Date</td><td>{domain_info['update_date']}</td></tr>")
                        
                        if domain_info.get('expiration_date'):
                            f.write(f"<tr><td>Expiration Date</td><td>{domain_info['expiration_date']}</td></tr>")
                        
                        # If no domain details were found, display a message
                        if not any(domain_info.get(field) for field in ['registrar', 'creation_date', 'update_date', 'expiration_date']):
                            f.write("<tr><td colspan='2'>No domain details available</td></tr>")
                        
                        f.write("</table>")
                        
                        # Domain Status
                        if domain_info.get('domain_status'):
                            f.write("<h3>Domain Status</h3>")
                            f.write("<ul>")
                            for status in domain_info['domain_status']:
                                f.write(f"<li>{status}</li>")
                            f.write("</ul>")
                        
                        # Name Servers
                        if domain_info.get('name_servers'):
                            f.write("<h3>Name Servers</h3>")
                            f.write("<ul>")
                            for ns in domain_info['name_servers']:
                                f.write(f"<li>{ns}</li>")
                            f.write("</ul>")
                        
                        f.write("</div>")

                        # MX Records - Add this section
                        if domain_info.get('mx_records'):
                            f.write("<h3>MX Records</h3>")
                            f.write("<ul>")
                            for mx in domain_info['mx_records']:
                                f.write(f"<li>{mx}</li>")
                            f.write("</ul>")
                        
                        f.write("</div>")  # End of domain info section

                        # SSL Certificate Information (right after domain info)
                        if self.target_url:
                            self._generate_ssl_certificate_section(f, self.target_url, domain_info)
                        
                        # IP Address Information
                        if domain_info.get('ip_addresses'):
                            f.write("<div class='section'>")
                            f.write("<h2>IP ADDRESS INFORMATION</h2>")
                            
                            for ip in domain_info['ip_addresses']:
                                f.write(f"<h3>{target_domain} -> {ip}</h3>")
                                
                                f.write("<table>")
                                if ip in self.ip_info:
                                    ip_data = self.ip_info[ip]
                                    if ip_data.get('cidr'):
                                        f.write(f"<tr><td>IP CIDR</td><td>{ip_data['cidr']}</td></tr>")
                                    
                                    if ip_data.get('asn'):
                                        asn_info = f"{ip_data['asn']}"
                                        if ip_data.get('organization'):
                                            asn_info += f" ({ip_data['organization']})"
                                        f.write(f"<tr><td>Origin AS</td><td>{asn_info}</td></tr>")
                                    
                                    if ip_data.get('country'):
                                        f.write(f"<tr><td>Country</td><td>{ip_data['country']}</td></tr>")
                                    
                                    if ip_data.get('reverse_dns'):
                                        f.write(f"<tr><td>Reverse DNS</td><td>{ip_data['reverse_dns']}</td></tr>")
                                else:
                                    f.write("<tr><td colspan='2'>No detailed IP information available</td></tr>")
                                f.write("</table>")
                            
                            f.write("</div>")
                    else:
                        f.write("<div class='section'>")
                        f.write("<h2>DOMAIN INFORMATION</h2>")
                        f.write(f"<p><strong>Domain:</strong> {target_domain}</p>")
                        f.write("<p>No WHOIS information could be retrieved for this domain.</p>")
                        f.write("</div>")
                
                # GPS Map section (if there are documents with GPS data)
                self._generate_gps_map_section(f)
                
                # Relationship Graph
                self._generate_relationship_graph(f)
                
                # Document Metadata section - now with collapsible sections by file type
                f.write("<div class='section'>")
                f.write("<h2>DOCUMENT METADATA INFORMATION</h2>")
                
                if self.document_metadata:
                    # Organize file types in a preferred order with friendly names
                    file_type_names = {
                        'pdf': 'PDF Documents',
                        'docx': 'Word Documents',
                        'xlsx': 'Excel Spreadsheets',
                        'pptx': 'PowerPoint Presentations',
                        'jpg': 'JPEG Images',
                        'jpeg': 'JPEG Images',
                        'png': 'PNG Images',
                        'gif': 'GIF Images',
                        'csv': 'CSV Files'
                    }
                    
                    # Sort file types by count (most documents first) and then alphabetically
                    sorted_file_types = sorted(
                        documents_by_type.keys(),
                        key=lambda x: (-len(documents_by_type[x]), x)
                    )
                    
                    # Create collapsible section for each file type
                    for file_type in sorted_file_types:
                        documents = documents_by_type[file_type]
                        display_name = file_type_names.get(file_type, f"{file_type.upper()} Files")
                        
                        # Create collapsible button for this document type
                        f.write(f'<button class="collapsible">{display_name} <span class="document-count">{len(documents)}</span></button>')
                        f.write(f'<div class="document-type-content">')
                        
                        # Write all documents of this type
                        for file_path, metadata in documents:
                            filename = os.path.basename(file_path)
                            
                            f.write(f"<div class='metadata-item'>")
                            f.write(f"<h3>{filename}</h3>")
                            f.write("<table>")
                            f.write(f"<tr><td>File Size</td><td>{metadata['file_size']} bytes</td></tr>")
                            
                            if metadata.get('title'):
                                f.write(f"<tr><td>Title</td><td>{metadata['title']}</td></tr>")
                            
                            if metadata.get('subject'):
                                f.write(f"<tr><td>Subject</td><td>{metadata['subject']}</td></tr>")
                            
                            if metadata.get('creation_date'):
                                f.write(f"<tr><td>Creation Date</td><td>{metadata['creation_date']}</td></tr>")
                            
                            if metadata.get('modification_date'):
                                f.write(f"<tr><td>Modification Date</td><td>{metadata['modification_date']}</td></tr>")
                            f.write("</table>")
                            
                            if metadata.get('authors'):
                                f.write("<h4>Authors/Users</h4>")
                                f.write("<ul>")
                                for author in sorted(metadata['authors']):
                                    f.write(f"<li>{author}</li>")
                                f.write("</ul>")
                            
                            if metadata.get('software'):
                                f.write("<h4>Software Used</h4>")
                                f.write("<ul>")
                                for sw in sorted(metadata['software']):
                                    f.write(f"<li>{sw}</li>")
                                f.write("</ul>")
                            
                            if metadata.get('found_emails'):
                                f.write("<h4>Emails Found in Document</h4>")
                                f.write("<ul>")
                                for email in sorted(metadata['found_emails']):
                                    f.write(f"<li>{email}</li>")
                                f.write("</ul>")
                            
                            if metadata.get('found_urls'):
                                f.write("<h4>URLs Found in Document</h4>")
                                f.write("<ul>")
                                for url in sorted(metadata['found_urls']):
                                    f.write(f"<li>{url}</li>")
                                f.write("</ul>")
                            
                            if metadata.get('found_paths'):
                                f.write("<h4>Paths Found in Document</h4>")
                                f.write("<ul>")
                                for path in sorted(metadata['found_paths']):
                                    f.write(f"<li>{path}</li>")
                                f.write("</ul>")
                            
                            # GPS data
                            if 'gps_data' in metadata and metadata['gps_data']:
                                f.write("<h4>GPS Coordinates</h4>")
                                f.write("<table>")
                                gps_data = metadata['gps_data']
                                if 'lat' in gps_data:
                                    f.write(f"<tr><td>Latitude</td><td>{gps_data['lat']}</td></tr>")
                                if 'lon' in gps_data:
                                    f.write(f"<tr><td>Longitude</td><td>{gps_data['lon']}</td></tr>")
                                if 'alt' in gps_data:
                                    f.write(f"<tr><td>Altitude</td><td>{gps_data['alt']}</td></tr>")
                                f.write("</table>")
                            
                            # Device info
                            if 'device_info' in metadata and metadata['device_info']:
                                f.write("<h4>Device Information</h4>")
                                f.write("<table>")
                                for key, value in metadata['device_info'].items():
                                    f.write(f"<tr><td>{key}</td><td>{value}</td></tr>")
                                f.write("</table>")
                            
                            # All Metadata Fields - FULL DETAILED LISTING
                            f.write("<h4>All Metadata Fields</h4>")
                            f.write("<table class='metadata-table'>")
                            f.write("<tr><th class='key-column'>Field</th><th class='value-column'>Value</th></tr>")
                            
                            if 'all_metadata' in metadata and metadata['all_metadata']:
                                # Sort keys for better readability
                                for key in sorted(metadata['all_metadata'].keys()):
                                    value = metadata['all_metadata'][key]
                                    if value is not None:
                                        # Format the value based on its type
                                        if isinstance(value, (list, dict)):
                                            formatted_value = json.dumps(value)
                                        else:
                                            formatted_value = str(value)
                                        f.write(f"<tr><td class='key-column'>{key}</td><td class='value-column'>{formatted_value}</td></tr>")
                            elif 'exiftool_metadata' in metadata and metadata['exiftool_metadata']:
                                # Flatten the nested metadata structure for display
                                flattened = self._flatten_metadata(metadata['exiftool_metadata'])
                                for key in sorted(flattened.keys()):
                                    value = flattened[key]
                                    if value is not None:
                                        # Format the value based on its type
                                        if isinstance(value, (list, dict)):
                                            formatted_value = json.dumps(value)
                                        else:
                                            formatted_value = str(value)
                                        f.write(f"<tr><td class='key-column'>{key}</td><td class='value-column'>{formatted_value}</td></tr>")
                            else:
                                f.write("<tr><td colspan='2'>No detailed metadata available</td></tr>")
                            
                            f.write("</table>")
                            f.write("</div>") # End of metadata item
                        
                        f.write("</div>") # End of collapsible content
                else:
                    f.write("<p>No document metadata found.</p>")
                
                f.write("</div>") # End of section
                
                # Footer
                f.write("""
                <div class="footer">
                    <p>Report generated by Sidikjari - Metadata Extraction Tool</p>
                    <p>Red Cell Security, LLC - www.redcellsecurity.org</p>
                </div>
            </div>
        </body>
        </html>""")
        except Exception as e:
            logger.error(f"Error generating HTML report: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())

    def _get_ssl_certificate_info(self, target_url):
        """Get SSL certificate information for a domain"""
        try:
            # Ensure target URL has a scheme
            if not target_url.startswith(('http://', 'https://')):
                target_url = f'https://{target_url}'
            
            parsed_url = urlparse(target_url)
            hostname = parsed_url.netloc
            
            # Remove port if present
            if ':' in hostname:
                hostname = hostname.split(':')[0]
            
            logger.info(f"Getting SSL certificate information for {hostname}")
            
            # Create a connection to get the certificate
            import ssl
            import socket
            import datetime
            
            context = ssl.create_default_context()
            with socket.create_connection((hostname, 443)) as sock:
                with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                    cert = ssock.getpeercert()
            
            # Process certificate information
            cert_info = {
                'subject': dict(x[0] for x in cert['subject']),
                'issuer': dict(x[0] for x in cert['issuer']),
                'version': cert['version'],
                'serialNumber': cert.get('serialNumber', 'N/A'),
                'notBefore': cert['notBefore'],
                'notAfter': cert['notAfter'],
                'subjectAltName': cert.get('subjectAltName', []),
                'OCSP': cert.get('OCSP', 'N/A'),
                'caIssuers': cert.get('caIssuers', 'N/A'),
                'crlDistributionPoints': cert.get('crlDistributionPoints', 'N/A')
            }
            
            # Format subject and issuer info
            cert_info['subject_str'] = ', '.join(f"{k}={v}" for k, v in cert_info['subject'].items())
            cert_info['issuer_str'] = ', '.join(f"{k}={v}" for k, v in cert_info['issuer'].items())
            
            # Format dates
            def parse_cert_date(date_str):
                return datetime.datetime.strptime(date_str, '%b %d %H:%M:%S %Y %Z')
            
            not_before = parse_cert_date(cert_info['notBefore'])
            not_after = parse_cert_date(cert_info['notAfter'])
            cert_info['valid_from'] = not_before.strftime('%Y-%m-%d %H:%M:%S')
            cert_info['valid_until'] = not_after.strftime('%Y-%m-%d %H:%M:%S')
            
            # Calculate validity period
            now = datetime.datetime.now()
            cert_info['is_valid'] = now >= not_before and now <= not_after
            cert_info['days_remaining'] = (not_after - now).days
            
            # Extract alternative names
            cert_info['alt_names'] = []
            for type_name, value in cert_info['subjectAltName']:
                if type_name == 'DNS':
                    cert_info['alt_names'].append(value)
            
            # Get certificate algorithm information
            if hasattr(ssock, 'cipher'):
                cipher = ssock.cipher()
                if cipher:
                    cert_info['cipher'] = cipher[0]
                    cert_info['protocol'] = cipher[1]
                    cert_info['secret_bits'] = cipher[2]
            
            # Check for certificate extensions
            cert_info['extensions'] = {}
            for oid, value in cert.get('extensions', []):
                cert_info['extensions'][oid] = value
            
            # Evaluate certificate strength
            # Initialize as secure, downgrade based on findings
            cert_info['security_assessment'] = 'Strong'
            cert_info['security_issues'] = []
            
            # Check expiration
            if cert_info['days_remaining'] < 30:
                cert_info['security_assessment'] = 'Warning'
                cert_info['security_issues'].append(f'Certificate expires soon ({cert_info["days_remaining"]} days remaining)')
            
            # Check if certificate is self-signed
            if cert_info['subject_str'] == cert_info['issuer_str']:
                cert_info['security_assessment'] = 'Weak'
                cert_info['security_issues'].append('Self-signed certificate')
            
            # Check for weak ciphers or protocols
            weak_protocols = ['SSLv2', 'SSLv3', 'TLSv1', 'TLSv1.1']
            if 'protocol' in cert_info and any(wp in cert_info['protocol'] for wp in weak_protocols):
                cert_info['security_assessment'] = 'Weak'
                cert_info['security_issues'].append(f'Weak protocol: {cert_info["protocol"]}')
            
            logger.info(f"Successfully retrieved SSL certificate information for {hostname}")
            return cert_info
        
        except Exception as e:
            logger.error(f"Error retrieving SSL certificate for {target_url}: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())
            return None

    def _generate_ssl_certificate_section(self, f, target_url, domain_info):
        """Generate a section with SSL certificate information in the HTML report"""
        try:
            # Get SSL certificate information
            cert_info = self._get_ssl_certificate_info(target_url)
            
            if not cert_info:
                return
            
            # Create SSL certificate section
            f.write("<div class='section'>")
            f.write("<h2>SSL CERTIFICATE INFORMATION</h2>")
            
            # Security assessment badge
            security_color = {
                'Strong': '#48BB78',  # Green
                'Warning': '#ECC94B',  # Yellow
                'Weak': '#F56565'     # Red
            }
            
            f.write(f"""
            <div style="margin-bottom: 20px;">
                <div style="display: inline-block; padding: 8px 16px; background-color: {security_color.get(cert_info['security_assessment'], '#718096')}; 
                     color: white; border-radius: 20px; font-weight: bold;">
                    Certificate Security: {cert_info['security_assessment']}
                </div>
            </div>
            """)
            
            # Certificate summary
            f.write("<table>")
            f.write(f"<tr><td width='180'><strong>Common Name</strong></td><td>{cert_info['subject'].get('commonName', 'N/A')}</td></tr>")
            f.write(f"<tr><td><strong>Issuer</strong></td><td>{cert_info['issuer'].get('organizationName', 'N/A')} {cert_info['issuer'].get('commonName', '')}</td></tr>")
            f.write(f"<tr><td><strong>Valid From</strong></td><td>{cert_info['valid_from']}</td></tr>")
            f.write(f"<tr><td><strong>Valid Until</strong></td><td>{cert_info['valid_until']} ({cert_info['days_remaining']} days remaining)</td></tr>")
            
            # Display protocol and cipher if available
            if 'protocol' in cert_info:
                f.write(f"<tr><td><strong>Protocol</strong></td><td>{cert_info['protocol']}</td></tr>")
            
            if 'cipher' in cert_info:
                f.write(f"<tr><td><strong>Cipher</strong></td><td>{cert_info['cipher']}</td></tr>")
            
            f.write("</table>")
            
            # Security issues if any
            if cert_info['security_issues']:
                f.write("<h3>Security Issues</h3>")
                f.write("<ul>")
                for issue in cert_info['security_issues']:
                    f.write(f"<li>{issue}</li>")
                f.write("</ul>")
            
            # Alternative names (SAN)
            if cert_info['alt_names']:
                f.write("<h3>Subject Alternative Names</h3>")
                f.write("<div style='max-height: 200px; overflow-y: auto; margin-bottom: 20px; padding: 10px; background-color: #f9f9f9; border-radius: 5px;'>")
                
                # Display in columns for better readability if many names
                if len(cert_info['alt_names']) > 5:
                    f.write("<div style='column-count: 2; column-gap: 20px;'>")
                else:
                    f.write("<div>")
                
                for name in cert_info['alt_names']:
                    f.write(f"<div style='margin-bottom: 5px;'>{name}</div>")
                
                f.write("</div></div>")
            
            # Certificate details (collapsible)
            f.write("""
            <button class="collapsible">View Full Certificate Details</button>
            <div class="document-type-content">
                <table class="metadata-table">
                    <tr><th class="key-column">Field</th><th class="value-column">Value</th></tr>
            """)
            
            # Display all certificate fields
            flattened_cert = self._flatten_metadata(cert_info)
            for key in sorted(flattened_cert.keys()):
                if key not in ['security_issues', 'alt_names']:  # Skip arrays already displayed
                    value = flattened_cert[key]
                    if value is not None:
                        # Format the value based on its type
                        if isinstance(value, (list, dict)):
                            formatted_value = json.dumps(value)
                        else:
                            formatted_value = str(value)
                        f.write(f"<tr><td class='key-column'>{key}</td><td class='value-column'>{formatted_value}</td></tr>")
            
            f.write("</table></div>")
            
            f.write("</div>")  # End of section
            
        except Exception as e:
            logger.error(f"Error generating SSL certificate section: {str(e)}")
            # Print traceback for debugging
            import traceback
            logger.error(traceback.format_exc())

    def _capture_website_screenshot(self, target_url):
        """Capture a screenshot of the target website's landing page using wkhtmltopdf"""
        try:
            # Create a directory for screenshots if it doesn't exist
            screenshots_dir = os.path.join(self.output_dir, "screenshots")
            os.makedirs(screenshots_dir, exist_ok=True)
            
            # Generate a filename for the screenshot
            domain = urlparse(target_url).netloc
            if not domain:
                domain = "website"
            screenshot_path = os.path.join(screenshots_dir, f"{domain}_screenshot.png")
            
            logger.info(f"Capturing screenshot of {target_url} using wkhtmltoimage")
            
            # Use wkhtmltoimage (part of wkhtmltopdf package)
            try:
                # Check if wkhtmltoimage is available
                if not shutil.which('wkhtmltoimage'):
                    logger.error("wkhtmltoimage tool not found. Please make sure wkhtmltopdf is installed correctly.")
                    return None
                    
                # Build the command
                cmd = [
                    'wkhtmltoimage',
                    '--width', '1366',       # Set width
                    '--height', '768',       # Set height
                    '--quality', '90',       # High quality
                    '--javascript-delay', '2000',  # Wait 2 seconds for JavaScript
                    '--no-stop-slow-scripts',      # Don't stop for slow scripts
                    '--disable-smart-width',       # Use specified width
                    '--enable-local-file-access',  # Allow local file access if needed
                    '--load-error-handling', 'ignore',  # Ignore load errors
                ]
                
                # Add user-agent to avoid bot detection
                cmd.extend(['--custom-header', 'User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'])
                
                # Add the URL and output path
                cmd.extend([target_url, screenshot_path])
                
                # Execute the command with a timeout
                process = subprocess.run(
                    cmd,
                    timeout=30,  # 30-second timeout
                    check=False, # Don't raise exception on non-zero exit
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE
                )
                
                # Check if the command was successful
                if process.returncode != 0:
                    logger.warning(f"wkhtmltoimage returned non-zero exit code: {process.returncode}")
                    logger.warning(f"Stderr: {process.stderr.decode('utf-8', errors='ignore')}")
                
                # Check if the file exists and has content
                if os.path.exists(screenshot_path) and os.path.getsize(screenshot_path) > 0:
                    logger.info(f"Screenshot saved to {screenshot_path}")
                    return screenshot_path
                else:
                    logger.warning(f"Screenshot file is empty or does not exist: {screenshot_path}")
                    
                    # Try with simpler options if the first attempt failed
                    simple_cmd = [
                        'wkhtmltoimage',
                        '--width', '1024',
                        '--height', '768',
                        '--disable-javascript',  # Disable JavaScript completely
                        target_url,
                        screenshot_path
                    ]
                    
                    process = subprocess.run(
                        simple_cmd,
                        timeout=20,
                        check=False,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE
                    )
                    
                    if os.path.exists(screenshot_path) and os.path.getsize(screenshot_path) > 0:
                        logger.info(f"Screenshot saved using simplified options to {screenshot_path}")
                        return screenshot_path
                    else:
                        logger.error("Both wkhtmltoimage attempts failed")
                        return None
                        
            except subprocess.TimeoutExpired:
                logger.error(f"Timeout while running wkhtmltoimage for {target_url}")
                return None
            except Exception as wk_e:
                logger.error(f"Error using wkhtmltoimage: {str(wk_e)}")
                return None
                
        except Exception as e:
            logger.error(f"Error in screenshot function: {str(e)}")
            return None
        
    def _generate_screenshot_section(self, f, target_url):
        """Generate a section with a screenshot of the website"""
        # Check if we already have a screenshot
        screenshots_dir = os.path.join(self.output_dir, "screenshots")
        domain = urlparse(target_url).netloc
        if not domain:
            domain = "website"
        screenshot_path = os.path.join(screenshots_dir, f"{domain}_screenshot.png")
        
        # Take screenshot if we don't have one yet
        if not os.path.exists(screenshot_path):
            screenshot_path = self._capture_website_screenshot(target_url)
        
        # If we have a screenshot, display it
        if screenshot_path and os.path.exists(screenshot_path):
            # Get the relative path for HTML embedding
            rel_path = os.path.relpath(screenshot_path, self.output_dir)
            
            # Create screenshot section
            f.write("<div class='section'>")
            f.write("<h2>WEBSITE SCREENSHOT</h2>")
            
            # Add timestamp - import datetime properly to avoid the previous error
            from datetime import datetime  # Make sure datetime is imported here
            capture_time = datetime.fromtimestamp(os.path.getmtime(screenshot_path))
            f.write(f"<p>Screenshot captured on: {capture_time.strftime('%Y-%m-%d %H:%M:%S')}</p>")
            
            # Display the screenshot with responsive sizing
            f.write(f"""
            <div style="text-align: center; margin: 20px 0;">
                <img src="{rel_path}" alt="Website Screenshot" style="max-width: 100%; border: 1px solid #ddd; border-radius: 5px; box-shadow: 0 2px 5px rgba(0,0,0,0.1);" />
            </div>
            """)
            
            # Add a link to open the full-size screenshot
            f.write(f"""
            <div style="text-align: center; margin-bottom: 20px;">
                <a href="{rel_path}" target="_blank" style="display: inline-block; padding: 8px 16px; background-color: #3498db; color: white; text-decoration: none; border-radius: 4px;">
                    View Full Size Screenshot
                </a>
            </div>
            """)
            
            f.write("</div>")  # End of section

    def _generate_gps_map_section(self, f):
        """Generate an interactive map section for documents with GPS coordinates"""
        # Collect GPS coordinates from all documents
        gps_locations = []
        for file_path, metadata in self.document_metadata.items():
            if 'gps_data' in metadata and metadata['gps_data']:
                gps_data = metadata['gps_data']
                if 'lat' in gps_data and 'lon' in gps_data:
                    try:
                        # Convert GPS coords to float if they're not already
                        lat = float(gps_data['lat']) if isinstance(gps_data['lat'], str) else gps_data['lat']
                        lon = float(gps_data['lon']) if isinstance(gps_data['lon'], str) else gps_data['lon']
                        
                        # Add to locations list with document info
                        gps_locations.append({
                            'lat': lat,
                            'lon': lon,
                            'filename': os.path.basename(file_path),
                            'filetype': metadata['file_type'],
                            'file_path': file_path
                        })
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid GPS coordinates in {file_path}: {gps_data}")
        
        # Only generate map if we have locations
        if not gps_locations:
            return
        
        # Write map section
        f.write("<div class='section'>")
        f.write("<h2>GPS COORDINATE MAP</h2>")
        
        # Table of GPS coordinates
        f.write("<table>")
        f.write("<tr><th>File</th><th>Type</th><th>Latitude</th><th>Longitude</th></tr>")
        
        for loc in gps_locations:
            f.write(f"<tr data-lat='{loc['lat']}' data-lon='{loc['lon']}' class='location-row' style='cursor:pointer;'>")
            f.write(f"<td>{loc['filename']}</td>")
            f.write(f"<td>{loc['filetype']}</td>")
            f.write(f"<td>{loc['lat']}</td>")
            f.write(f"<td>{loc['lon']}</td>")
            f.write("</tr>")
        
        f.write("</table>")
        
        # Map container
        f.write("<div id='map' style='height: 500px; margin-top: 20px;'></div>")
        
        # Leaflet.js library and custom script
        f.write("""
        <link rel="stylesheet" href="https://unpkg.com/leaflet@1.7.1/dist/leaflet.css" />
        <script src="https://unpkg.com/leaflet@1.7.1/dist/leaflet.js"></script>
        <script>
            document.addEventListener('DOMContentLoaded', function() {
                // Initialize map
                var map = L.map('map');
                
                // Add OpenStreetMap tile layer
                L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
                    attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors'
                }).addTo(map);
                
                // Define location data
                var locations = [
        """)
        
        # Add location data as JavaScript array
        for loc in gps_locations:
            f.write(f"""                {{
                        lat: {loc['lat']},
                        lon: {loc['lon']},
                        name: "{loc['filename']}",
                        type: "{loc['filetype']}"
                    }},
        """)
        
        f.write("""            ];
                
                // Create markers for each location
                var markers = [];
                locations.forEach(function(loc) {
                    var marker = L.marker([loc.lat, loc.lon])
                        .addTo(map)
                        .bindPopup("<b>" + loc.name + "</b><br>Type: " + loc.type + "<br>Coordinates: " + loc.lat + ", " + loc.lon);
                    markers.push(marker);
                });
                
                // Set view to fit all markers
                if (markers.length > 0) {
                    var group = new L.featureGroup(markers);
                    map.fitBounds(group.getBounds().pad(0.1));
                } else {
                    map.setView([0, 0], 2); // Default view if no markers
                }
                
                // Add click handler for table rows
                document.querySelectorAll('.location-row').forEach(function(row) {
                    row.addEventListener('click', function() {
                        var lat = parseFloat(this.getAttribute('data-lat'));
                        var lon = parseFloat(this.getAttribute('data-lon'));
                        map.setView([lat, lon], 15);
                        
                        // Find and open the corresponding marker popup
                        markers.forEach(function(marker) {
                            var markerLatLng = marker.getLatLng();
                            if (markerLatLng.lat === lat && markerLatLng.lng === lon) {
                                marker.openPopup();
                            }
                        });
                    });
                });
            });
        </script>
        """)
        
        f.write("</div>") # End of section

    def _generate_relationship_graph(self, f):
        """Generate an interactive social graph visualization of entity relationships"""
        # Skip if there's no useful data to visualize
        if not self.users and not self.emails and not self.internal_domains:
            return
        
        # Prepare nodes and links data
        nodes = []
        links = []
        node_index = {}  # To track node indices
        
        # Helper to add a node if it doesn't exist
        def add_node(id, label, type):
            if id not in node_index:
                node_index[id] = len(nodes)
                nodes.append({
                    "id": id,
                    "label": label,
                    "type": type
                })
            return node_index[id]
        
        # Filter out non-human emails
        system_emails = ['noreply', 'no-reply', 'donotreply', 'admin@', 'administrator', 
                         'info@', 'support@', 'help@', 'contact@', 'webmaster@', 
                         'postmaster@', 'hostmaster@', 'sales@', 'marketing@']
        
        human_emails = []
        for email in self.emails:
            if not any(pattern in email.lower() for pattern in system_emails):
                human_emails.append(email)
        
        # Add users first
        for user in self.users:
            # Skip long strings that are likely not user names
            if len(user) > 40 or user.startswith('/') or '\\' in user:
                continue
            add_node(f"user_{user}", user, "user")
        
        # Add emails and create links to users
        for email in human_emails:
            if '@' in email:
                username, domain = email.split('@')
                
                # Skip technical or system usernames
                if username.lower() in ['administrator', 'admin', 'support', 'info', 'contact']:
                    continue
                    
                # Add email node
                email_idx = add_node(f"email_{email}", email, "email")
                
                # Add domain node
                domain_idx = add_node(f"domain_{domain}", domain, "domain")
                
                # Link email to domain
                links.append({
                    "source": email_idx,
                    "target": domain_idx,
                    "type": "belongs_to"
                })
                
                # Link users to emails if username matches
                for user in self.users:
                    # Skip non-user strings
                    if len(user) > 40 or user.startswith('/') or '\\' in user:
                        continue
                    
                    # Simple matching - can be improved
                    if (user.lower() in username.lower() or 
                        username.lower() in user.lower() or
                        self._calculate_similarity(user.lower(), username.lower()) > 0.7):
                        user_idx = node_index.get(f"user_{user}")
                        if user_idx is not None:  # Ensure the user node exists
                            links.append({
                                "source": user_idx,
                                "target": email_idx,
                                "type": "owns"
                            })
        
        # Add domains and their relationships
        for domain in self.internal_domains:
            # Skip overly long domains that might be error text
            if len(domain) > 50:
                continue
                
            domain_idx = add_node(f"domain_{domain}", domain, "domain")
            
            # Link domains to IPs
            for ip in self.ip_addresses:
                # Skip non-IP strings that might have been wrongly classified
                try:
                    ipaddress.ip_address(ip)  # Validate IP format
                    if ip in self.ip_info and domain in self.ip_info[ip].get('associated_domains', []):
                        ip_idx = add_node(f"ip_{ip}", ip, "ip")
                        links.append({
                            "source": domain_idx,
                            "target": ip_idx,
                            "type": "resolves_to"
                        })
                except ValueError:
                    continue
        
        # Generate HTML for the visualization only if we have meaningful data
        if len(nodes) > 1 and len(links) > 0:
            f.write("<div class='section'>")
            f.write("<h2>RELATIONSHIP GRAPH</h2>")
            f.write("<p>Interactive visualization of relationships between entities discovered in metadata.</p>")
            
            # Controls for the graph
            f.write("""
            <div style="margin-bottom: 15px;">
                <div style="margin-bottom: 10px;">
                    <strong>Filter by type:</strong>
                    <label><input type="checkbox" class="node-type" value="user" checked> Users</label>
                    <label><input type="checkbox" class="node-type" value="email" checked> Emails</label>
                    <label><input type="checkbox" class="node-type" value="domain" checked> Domains</label>
                    <label><input type="checkbox" class="node-type" value="ip" checked> IP Addresses</label>
                </div>
                <button id="reset-zoom" style="margin-right: 10px;">Reset Zoom</button>
                <input type="range" id="link-distance" min="30" max="300" value="100">
                <label for="link-distance">Link Distance</label>
            </div>
            """)
            
            # SVG container for the graph
            f.write('<svg id="relationship-graph" width="100%" height="600" style="border: 1px solid #ccc; border-radius: 5px;"></svg>')
            
            # Load D3.js and add visualization code
            f.write("""
            <script src="https://d3js.org/d3.v7.min.js"></script>
            <script>
            document.addEventListener('DOMContentLoaded', function() {
                // Graph data
                const nodes = """)
            f.write(json.dumps(nodes))
            f.write(";\n        const links = ")
            f.write(json.dumps(links))
            f.write(""";
                
                // Node colors by type
                const colors = {
                    user: "#4299E1",   // Blue
                    email: "#48BB78",  // Green
                    domain: "#ED8936", // Orange
                    ip: "#9F7AEA"      // Purple
                };
                
                // Node sizes by type
                const sizes = {
                    user: 8,
                    email: 6,
                    domain: 10,
                    ip: 7
                };
                
                // Set up the SVG
                const svg = d3.select("#relationship-graph");
                const width = svg.node().getBoundingClientRect().width;
                const height = svg.node().getBoundingClientRect().height;
                
                // Create zoom behavior
                const zoom = d3.zoom()
                    .scaleExtent([0.1, 4])
                    .on("zoom", (event) => {
                        g.attr("transform", event.transform);
                    });
                
                svg.call(zoom);
                
                // Create container for the graph
                const g = svg.append("g");
                
                // Initialize the simulation
                const simulation = d3.forceSimulation(nodes)
                    .force("link", d3.forceLink(links).id(d => d.id).distance(100))
                    .force("charge", d3.forceManyBody().strength(-200))
                    .force("center", d3.forceCenter(width / 2, height / 2))
                    .force("collision", d3.forceCollide().radius(d => sizes[d.type] * 2));
                
                // Create links
                const link = g.append("g")
                    .selectAll("line")
                    .data(links)
                    .join("line")
                    .attr("stroke", "#999")
                    .attr("stroke-opacity", 0.6)
                    .attr("stroke-width", 1);
                
                // Create nodes
                const node = g.append("g")
                    .selectAll("circle")
                    .data(nodes)
                    .join("circle")
                    .attr("r", d => sizes[d.type])
                    .attr("fill", d => colors[d.type])
                    .attr("class", d => `node-${d.type}`)
                    .call(drag(simulation));
                
                // Add labels
                const label = g.append("g")
                    .selectAll("text")
                    .data(nodes)
                    .join("text")
                    .text(d => d.label)
                    .attr("font-size", 8)
                    .attr("dx", 12)
                    .attr("dy", ".35em")
                    .attr("class", d => `label-${d.type}`)
                    .attr("pointer-events", "none");
                
                // Initialize the simulation
                simulation.on("tick", () => {
                    link
                        .attr("x1", d => d.source.x)
                        .attr("y1", d => d.source.y)
                        .attr("x2", d => d.target.x)
                        .attr("y2", d => d.target.y);
                    
                    node
                        .attr("cx", d => d.x)
                        .attr("cy", d => d.y);
                    
                    label
                        .attr("x", d => d.x)
                        .attr("y", d => d.y);
                });
                
                // Add tooltips
                node.append("title")
                    .text(d => `${d.label} (${d.type})`);
                
                // Implement dragging
                function drag(simulation) {
                    function dragstarted(event) {
                        if (!event.active) simulation.alphaTarget(0.3).restart();
                        event.subject.fx = event.subject.x;
                        event.subject.fy = event.subject.y;
                    }
                    
                    function dragged(event) {
                        event.subject.fx = event.x;
                        event.subject.fy = event.y;
                    }
                    
                    function dragended(event) {
                        if (!event.active) simulation.alphaTarget(0);
                        event.subject.fx = null;
                        event.subject.fy = null;
                    }
                    
                    return d3.drag()
                        .on("start", dragstarted)
                        .on("drag", dragged)
                        .on("end", dragended);
                }
                
                // Handle type filtering
                document.querySelectorAll('.node-type').forEach(checkbox => {
                    checkbox.addEventListener('change', function() {
                        const type = this.value;
                        const isChecked = this.checked;
                        
                        // Update node visibility
                        node.filter(d => d.type === type)
                            .style("display", isChecked ? "block" : "none");
                        
                        // Update label visibility
                        label.filter(d => d.type === type)
                            .style("display", isChecked ? "block" : "none");
                        
                        // Update link visibility
                        link.style("display", function(d) {
                            const sourceType = nodes[links.indexOf(d)].source.type;
                            const targetType = nodes[links.indexOf(d)].target.type;
                            
                            // Check if either end of the link is hidden
                            const sourceVisible = document.querySelector(`.node-type[value="${sourceType}"]`).checked;
                            const targetVisible = document.querySelector(`.node-type[value="${targetType}"]`).checked;
                            
                            return (sourceVisible && targetVisible) ? "block" : "none";
                        });
                        
                        // Reheat the simulation
                        simulation.alpha(0.3).restart();
                    });
                });
                
                // Reset zoom
                document.getElementById('reset-zoom').addEventListener('click', function() {
                    svg.transition().duration(750).call(
                        zoom.transform,
                        d3.zoomIdentity,
                        d3.zoomTransform(svg.node()).invert([width / 2, height / 2])
                    );
                });
                
                // Update link distance
                document.getElementById('link-distance').addEventListener('input', function() {
                    const distance = parseInt(this.value);
                    simulation.force("link").distance(distance);
                    simulation.alpha(0.3).restart();
                });
            });
            </script>
            """)
            
            f.write("</div>") # End of section

    def _calculate_similarity(self, str1, str2):
        """Calculate string similarity ratio for fuzzy matching"""
        from difflib import SequenceMatcher
        return SequenceMatcher(None, str1, str2).ratio()

    def run(self):
        """Execute the full analysis"""
        logger.info(f"{Fore.GREEN}Starting Sidikjari analysis on {self.target_url}{Style.RESET_ALL}")
        
        # Step 1: Crawl website
        self.crawl_website()
        
        # Step 2: Download documents
        self.download_documents()
        
        # Step 3: Extract metadata
        self.extract_all_metadata()
        
        # Step 4: Generate the report
        report_path = self.generate_reports()
        
        logger.info(f"{Fore.GREEN}Analysis complete! Report available at: {report_path}{Style.RESET_ALL}")
        self._print_summary()
    
    def _print_summary(self):
        """Print a summary of the findings"""
        # Make sure to use Rich's Table class, not ReportLab's
        from rich.table import Table as RichTable
        
        table = RichTable(title="Sidikjari Analysis Summary")
        table.add_column("Item")
        table.add_column("Count")
        
        table.add_row("Documents Analyzed", str(len(self.file_paths)))
        table.add_row("Users/Authors Discovered", str(len(self.users)))
        table.add_row("Email Addresses Discovered", str(len(self.emails)))
        table.add_row("Software Identified", str(len(self.software)))
        table.add_row("Domains Discovered", str(len(self.internal_domains)))
        table.add_row("IP Addresses Found", str(len(self.ip_addresses)))
        
        console.print(table)

def main():
    parser = argparse.ArgumentParser(description="Sidikjari - Metadata extraction and analysis tool")
    parser.add_argument("--url", "-u", help="Target URL to scan")
    parser.add_argument("--output", "-o", default="output", help="Output directory")
    parser.add_argument("--depth", "-d", type=int, default=2, 
                       help="Crawl depth (0=homepage only, 1=direct links, 2=links from direct links, etc.). Higher values crawl more of the site but take longer.")
    parser.add_argument("--threads", "-t", type=int, default=10, help="Number of threads")
    parser.add_argument("--local", "-l", help="Local directory of files to analyze (instead of URL)")
    parser.add_argument("--time-delay", type=float, default=0.0,
                       help="Delay in seconds between web requests to avoid overwhelming the server (e.g., 0.5)")
    parser.add_argument("--user-agent", choices=["default", "chrome", "firefox", "safari", "edge", "mobile", "random"],
                       default="default", help="User agent to use for web requests")
    
    args = parser.parse_args()
    
    # Create output directory before setting up logging
    output_dir = args.output
    os.makedirs(output_dir, exist_ok=True)
    
    # Reconfigure logging to use the output directory
    log_path = os.path.join(output_dir, "Sidikjari.log")
    
    # Clear existing handlers and add new ones
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_path),
            logging.StreamHandler()
        ]
    )
    
    logger = logging.getLogger("Sidikjari")
    
    # Print banner
    print(f"""{Fore.CYAN}
    ████████████████████████
    █                                           █
    █              SIDIKJARI                    █
    █                                           █
    █  Metadata extraction and analysis tool    █
    █                                           █
    ████████████████████████
    {Style.RESET_ALL}""")
    
    if not args.url and not args.local:
        parser.print_help()
        print(f"\n{Fore.RED}Error: You must provide either a URL to scan or a local directory{Style.RESET_ALL}")
        sys.exit(1)
    
    # Check for required dependencies
    if not shutil.which('exiftool'):
        print(f"\n{Fore.RED}Error: ExifTool is required but not found in PATH. Please install ExifTool.{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Installation instructions: https://exiftool.org/install.html{Style.RESET_ALL}")
        sys.exit(1)
    
    # Check for Selenium if using URL scanning (for screenshots)
    if args.url:
        try:
            from selenium import webdriver
        except ImportError:
            print(f"\n{Fore.YELLOW}Warning: Selenium is not installed. Website screenshots will be disabled.{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}To _analyze_domain_infoenable screenshots, install Selenium: pip install selenium webdriver-manager{Style.RESET_ALL}")
    
    try:
        if args.url:
            # URL-based scanning
            target_url = args.url
            if not target_url.startswith(('http://', 'https://')):
                target_url = f'https://{target_url}'
                
            sidikjari_scanner = Sidikjari(
                target_url=target_url,
                output_dir=args.output,
                depth=args.depth,
                threads=args.threads,
                time_delay=args.time_delay,
                user_agent=args.user_agent
            )
            
            # Run the full analysis
            sidikjari_scanner.run()
            
        else:
            # Local directory scanning
            print(f"{Fore.GREEN}Analyzing local directory: {args.local}{Style.RESET_ALL}")

            local_sidikjari = LocalSidikjari(
                input_dir=args.local,
                output_dir=args.output,
                threads=args.threads
            )
            
            # Run the full analysis
            local_sidikjari.run()
            
    except Exception as e:
        logger.error(f"Error during execution: {str(e)}")
        print(f"\n{Fore.RED}An error occurred: {str(e)}{Style.RESET_ALL}")
        # Print traceback for debugging
        import traceback
        traceback.print_exc()
        sys.exit(1)

class LocalSidikjari(Sidikjari):
    """Version of Sidikjari that works with local files instead of crawling websites"""
    
    def __init__(self, input_dir, output_dir="output", threads=10, report_format="text"):
        super().__init__(target_url=None, output_dir=output_dir, threads=threads, report_format=report_format)
        self.input_dir = input_dir
    
    def crawl_website(self):
        """Find local files instead of crawling websites"""
        logger.info(f"{Fore.GREEN}Scanning directory: {self.input_dir}{Style.RESET_ALL}")
        self._find_local_documents()
    
    def download_documents(self):
        """No need to download documents as they're already local"""
        logger.info(f"{Fore.GREEN}Found {len(self.file_paths)} local documents to analyze{Style.RESET_ALL}")
        
    def _find_local_documents(self):
        """Find all documents in the input directory"""
        for root, _, files in os.walk(self.input_dir):
            for file in files:
                file_path = os.path.join(root, file)
                file_ext = os.path.splitext(file)[1].lower().replace('.', '')
                
                if file_ext in self.interesting_extensions:
                    self.file_paths.add(file_path)
                    logger.info(f"Found document to analyze: {file_path} ({file_ext})")
        
        logger.info(f"{Fore.GREEN}Found {len(self.file_paths)} documents{Style.RESET_ALL}")

if __name__ == "__main__":
    main()
